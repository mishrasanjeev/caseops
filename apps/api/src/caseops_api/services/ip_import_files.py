"""Bounded CSV/XLSX parsing for the neutral IP import owner."""

from __future__ import annotations

import csv
import io
from pathlib import Path

from fastapi import HTTPException, status
from openpyxl import load_workbook

from caseops_api.schemas.ip_imports import IpImportJobCreateRequest, IpImportRowInput

MAX_IMPORT_BYTES = 10 * 1024 * 1024
MAX_IMPORT_ROWS = 1000
HEADER_ALIASES = {
    "title": "title",
    "mark": "mark_text",
    "mark text": "mark_text",
    "mark_text": "mark_text",
    "class": "class_number",
    "nice class": "class_number",
    "class_number": "class_number",
    "applicant": "applicant_name",
    "proprietor": "applicant_name",
    "applicant_name": "applicant_name",
    "goods/services": "specification",
    "goods and services": "specification",
    "specification": "specification",
    "application number": "application_number",
    "application_number": "application_number",
    "representation": "representation_kind",
    "representation type": "representation_kind",
    "representation_kind": "representation_kind",
    "agent": "agent_name",
    "agent name": "agent_name",
    "agent_name": "agent_name",
    "jurisdiction": "jurisdiction",
    "office": "office",
    "matter id": "matter_id",
    "matter_id": "matter_id",
}


def _normalize_headers(values: list[object]) -> list[str]:
    normalized: list[str] = []
    for value in values:
        key = str(value or "").strip().casefold()
        normalized.append(HEADER_ALIASES.get(key, key.replace(" ", "_")))
    if len(set(normalized)) != len(normalized):
        raise HTTPException(status_code=422, detail="Import headers must be unique.")
    return normalized


def _rows_from_csv(content: bytes) -> list[dict[str, object]]:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise HTTPException(
            status_code=422,
            detail="CSV files must use UTF-8 encoding.",
        ) from exc
    reader = csv.reader(io.StringIO(text, newline=""))
    materialized = list(reader)
    if not materialized:
        raise HTTPException(status_code=422, detail="The import file is empty.")
    headers = _normalize_headers(list(materialized[0]))
    return [dict(zip(headers, values, strict=False)) for values in materialized[1:]]


def _rows_from_xlsx(content: bytes) -> list[dict[str, object]]:
    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=422, detail="The XLSX file could not be read.") from exc
    try:
        sheet = workbook.active
        iterator = sheet.iter_rows(values_only=True)
        headers = _normalize_headers(list(next(iterator)))
        return [dict(zip(headers, values, strict=False)) for values in iterator]
    except StopIteration as exc:
        raise HTTPException(status_code=422, detail="The import file is empty.") from exc
    finally:
        workbook.close()


def parse_ip_import_file(*, filename: str, content: bytes) -> IpImportJobCreateRequest:
    if not content:
        raise HTTPException(status_code=422, detail="The import file is empty.")
    if len(content) > MAX_IMPORT_BYTES:
        raise HTTPException(
            status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
            detail="Import files are limited to 10 MB.",
        )
    suffix = Path(filename).suffix.casefold()
    if suffix == ".csv":
        rows = _rows_from_csv(content)
    elif suffix == ".xlsx":
        rows = _rows_from_xlsx(content)
    else:
        raise HTTPException(status_code=415, detail="Use a CSV or XLSX import file.")
    rows = [row for row in rows if any(value not in (None, "") for value in row.values())]
    if not rows:
        raise HTTPException(status_code=422, detail="The import file has no data rows.")
    if len(rows) > MAX_IMPORT_ROWS:
        raise HTTPException(status_code=422, detail="Import files are limited to 1,000 rows.")
    return IpImportJobCreateRequest(
        filename=filename,
        rows=[
            IpImportRowInput(row_number=index, values=row)
            for index, row in enumerate(rows, start=1)
        ],
    )


__all__ = ["parse_ip_import_file"]
