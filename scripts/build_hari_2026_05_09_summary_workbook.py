"""
Builds the Hari 2026-05-09 sweep summary workbook for the user's
Downloads folder. One-shot script — re-running it re-generates the
file from scratch (deterministic output).

Sheets:
  - Summary
  - Bug Details
  - Verification Matrix
  - Reopen Learnings
  - Pending Prod Proof
"""
from __future__ import annotations

import os
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

OUT_PATH = os.path.expanduser(
    "~/Downloads/CaseOps Bug Fix Summary_Hari9May2026.xlsx"
)

HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(bold=True, color="FFFFFF")
WRAP = Alignment(wrap_text=True, vertical="top")
TOP = Alignment(vertical="top")


def _style_header_row(ws, row_idx: int, col_count: int) -> None:
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP


def _autosize(ws, min_w: int = 12, max_w: int = 60) -> None:
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        longest = 0
        for cell in col:
            v = "" if cell.value is None else str(cell.value)
            for line in v.splitlines() or [""]:
                longest = max(longest, len(line))
        ws.column_dimensions[col_letter].width = max(min_w, min(max_w, longest + 2))


def _add_header(ws, headers: list[str]) -> None:
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))
    ws.freeze_panes = "A2"


def _wrap_rows(ws, start_row: int = 2) -> None:
    for row in ws.iter_rows(min_row=start_row):
        for cell in row:
            cell.alignment = WRAP


def build_summary_sheet(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "Summary"
    ws.append(["CaseOps — Hari 2026-05-09 Bug Fix Sweep Summary"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    rows = [
        ["Source workbook", "C:\\Users\\mishr\\Downloads\\CaseOps Bug List_Hari9May2026 .xlsx"],
        ["Sweep date", "2026-05-09"],
        ["Closeout date", "2026-05-10"],
        ["Total areas worked", "5 (BUG-032, BUG-033, BUG-034, BUG-038, BUG-039)"],
        ["Total PRs opened", "5 (#20, #21, #22, #23, #24)"],
        ["Verdict status (per bug-fixing skill)", "All five Partially fixed pending merge + deploy + prod-Playwright"],
        ["Closeout PR", "Pending — this workbook + docs updates ship as the audit-summary PR"],
        [],
        ["Why every verdict is Partially fixed", "Per .claude/skills/bug-fixing/SKILL.md: a bug is Properly fixed only when the user can complete the intended workflow on the deployed surface, proven by a prod-Playwright spec passing on the deployed commit SHA. Pre-merge, none of the five PRs have that proof yet."],
        ["Key cross-PR risk", "playwright.app.config.ts and playwright.prod-ram.config.ts both gain testMatch entries from PRs #20, #21, #23, #24. Whichever merges first, the others must be rebased keeping ALL spec entries — dropping any silently disables that bug's prod regression."],
        ["Per-bug detail", "See 'Bug Details' sheet."],
        ["Verification commands run", "See 'Verification Matrix' sheet."],
        ["Brutal learnings", "See 'Reopen Learnings' sheet — also docs/BUG_REOPEN_LEARNINGS_2026-05-09.md."],
        ["What's left for Properly Fixed", "See 'Pending Prod Proof' sheet."],
    ]
    for r in rows:
        ws.append(r)
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
        for cell in row:
            cell.alignment = WRAP
        row[0].font = Font(bold=True)
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 110


def build_bug_details_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Bug Details")
    headers = [
        "Bug/Area ID",
        "Source",
        "Severity",
        "Module",
        "Classification",
        "Root Cause",
        "Fix Summary",
        "Adjacent Paths Audited",
        "Backend Tests",
        "Frontend Tests",
        "Playwright Spec",
        "Environment",
        "Commit SHA",
        "PR URL",
        "Verdict",
        "Caveats / Blockers",
        "Reopen Learning",
    ]
    _add_header(ws, headers)

    rows = [
        [
            "BUG-032",
            "Hari 2026-05-09 workbook",
            "P1",
            "Matter / Hearings / Documents",
            "Workflow gap (linked-record selector with no create path)",
            "MatterCourtOrder rows could only come from court-sync. Documents-page Linked-order selector was empty for any matter without a court-sync. No manual create path existed in API or UI.",
            "Added POST /api/matters/{id}/court-orders endpoint with MatterCourtOrderCreateRequest schema. Frontend AddCourtOrderDialog mounted in Hearings page header AND empty state. File-upload integrates the existing uploadMatterAttachment ladder so the same dialog produces both metadata-only and attachment-linked orders. Workspace cache invalidates on success so the documents-page Linked-order selector refreshes immediately.",
            "Documents-page LinkedOrderSelect (consumes the same workspace.court_orders array — verified picks up new orders).",
            "test_matter_court_orders_create.py — 7 tests: metadata-only create, with-attachment links, cross-matter rejected 400, cross-tenant rejected 404, audit + activity recorded, notification fires on attachment-linked, 422 on missing required fields.",
            "page.test.tsx — added BUG-032 case asserting ≥2 add-court-order-open triggers (header + empty state).",
            "tests/e2e/hari-2026-05-09-bug-032.spec.ts (local: API create surfaces in workspace; UI mounts ≥2 triggers). tests/e2e/hari-2026-05-09-bug-032-prod.spec.ts (prod: disposable matter + order on QA tenant; matters-list h1 renders).",
            "Local + prod-Playwright (prod run pending deploy of merged SHA).",
            "Per PR #24 head — see 'Verification Matrix' sheet for SHA.",
            "https://github.com/mishrasanjeev/caseops/pull/24",
            "Partially fixed pending merge + deploy + prod-Playwright run on deployed SHA",
            "testMatch overlap on playwright.app.config.ts + playwright.prod-ram.config.ts with PRs #20, #21, #23 — second-mover must rebase carrying ALL spec entries.",
            "L4 — linked-record metadata existed without a complete create/upload workflow.",
        ],
        [
            "BUG-033",
            "Hari 2026-05-09 workbook",
            "P1",
            "Auth / Account / Employee onboarding",
            "Missing Next.js routes for backend-generated email links",
            "/account/setup?token=... and /account/reset-password?token=... were generated by services/employee_mailer.py since LW-S5 employee admin shipped, but neither Next.js route existed on the frontend. Real users got the 404 'this page isn't on the matter graph' surface.",
            "Added apps/web/app/account/setup/page.tsx + reset-password/page.tsx using server-component wrapper + Suspense around the client AccountSetupForm/ResetPasswordForm (required for useSearchParams under app router). Added apps/web/app/account/layout.tsx wrapping in AppProviders (QueryClientProvider + Toaster).",
            "Verified mailer link generation now matches a real route. Verified employee_mailer.py + portal_mailer.py bypass the new EmailSuppression check (auth-flow mailers must never be silently suppressed) — regression test landed in PR #22.",
            "n/a (frontend route).",
            "page.test.tsx for AccountSetupForm + ResetPasswordForm — happy path + invalid-token + missing-token states.",
            "tests/e2e/hari-2026-05-09-bug-033.spec.ts (local: route renders without 404; form mounts). tests/e2e/hari-2026-05-09-bug-033-prod.spec.ts (prod: route renders on caseops.ai). Note: full happy-path with debug_token blocked because DEBUG_TOKEN_ENVS = {'local','test'} excludes 'e2e' — vitest + pytest carry the workflow proof.",
            "Local + prod-Playwright (prod run pending deploy of merged SHA).",
            "Per PR #21 head — see 'Verification Matrix' sheet for SHA.",
            "https://github.com/mishrasanjeev/caseops/pull/21",
            "Partially fixed pending merge + deploy + prod-Playwright run on deployed SHA",
            "Happy-path E2E gated on broadening DEBUG_TOKEN_ENVS or persisting a real token through the prod-QA tenant — not done in this PR. testMatch overlap on Playwright configs.",
            "L2 — email links generated for routes that did not exist.",
        ],
        [
            "BUG-034",
            "Hari 2026-05-09 workbook",
            "P1",
            "Teams / Roles / Capabilities",
            "Backend authorization rule not represented in UI controls",
            "services/capabilities.py::_NON_DELEGABLE_CUSTOM_ROLE_CAPABILITIES rejected email_templates:manage / portal:invite / portal:manage_grants for custom roles with 403, but the CapabilityRecord schema only carried owner_only:bool — frontend could only gate owner-only caps. Custom-role create form rendered protected caps as selectable, user discovered the rule by 403-on-submit.",
            "Added custom_role_delegable:bool + protected_reason:str|None to CapabilityRecord schema. Backend computes from the existing non-delegable set. Frontend disables non-delegable checkboxes and renders the protected_reason verbatim before submit.",
            "Audited owner-only path (still gated correctly). Audited every other consumer of the capability catalog endpoint (/teams/admin/roles, /teams/admin/users) — both render disabled + reason now.",
            "test_capabilities_catalog.py — protected_reason populated for non-delegable; delegable=true for normal caps; legacy owner_only behavior preserved. test_custom_roles_create_protected_capability.py — backend still rejects with 403 if frontend gating is bypassed.",
            "RoleCapabilityCheckbox.test.tsx — disabled state when custom_role_delegable=false; protected_reason rendered; submit-disabled when any selected cap is non-delegable.",
            "tests/e2e/hari-2026-05-09-bugs.spec.ts (local: catalog reports custom_role_delegable; UI disables protected caps before submit). tests/e2e/hari-2026-05-09-prod.spec.ts (prod: catalog endpoint reports the new fields).",
            "Local + prod-Playwright (prod run pending deploy of merged SHA).",
            "Per PR #20 head — see 'Verification Matrix' sheet for SHA.",
            "https://github.com/mishrasanjeev/caseops/pull/20",
            "Partially fixed pending merge + deploy + prod-Playwright run on deployed SHA",
            "testMatch overlap on Playwright configs (rebase guard).",
            "L3 — backend authorization rules not represented in UI controls.",
        ],
        [
            "BUG-038 (SendGrid)",
            "Operational drift / EH-PROV-01",
            "P1",
            "Email / Deliverability / Provider integration",
            "Provider integration code-complete while runtime/provider config incomplete",
            "SendGrid send had been working from imperatively-set Cloud Run env vars; the webhook side had no declarative wiring at all — no signature verification, no suppression table, no manifest secret reference, no runbook, no SendGrid dashboard config. Bounces / unsubscribes / spam reports were being silently dropped.",
            "ECDSA P-256 signed-event verification with explicit fail-closed on missing/invalid signature. Tenant-scoped EmailSuppression table ((company_id, recipient_email) unique). EmailSuppressionReason StrEnum. Idempotent on duplicate events. services/email_send.py calls is_suppressed before every business mailer; employee_mailer.py + portal_mailer.py intentionally bypass (auth flow can never be suppressed). Cloud Run manifest references caseops-sendgrid-webhook-public-key Secret Manager value. Runbook docs/runbooks/sendgrid-event-webhook.md.",
            "Audited every services/*_mailer.py call site for is_suppressed wiring. Confirmed employee_mailer + portal_mailer bypass — added test_auth_flow_mailers_bypass_suppression as regression lock.",
            "12 backend tests — valid-sig accept, invalid-sig 503 fail-closed, missing-sig 503, replay protection, tenant-scoped suppression isolation, group_unsubscribe handling, idempotent on duplicate, downgrade migration safe, auth-flow bypass, etc.",
            "n/a (server-only feature).",
            "n/a (no Playwright surface — server-only ingestion).",
            "Server-only; verdict ceiling depends on operator-side SendGrid dashboard + Secret Manager config.",
            "Per PR #22 head — see 'Verification Matrix' sheet for SHA.",
            "https://github.com/mishrasanjeev/caseops/pull/22",
            "Partially fixed pending merge + deploy + operator-side runbook completion",
            "BLOCKER: Secret Manager value caseops-sendgrid-webhook-public-key must be created with SendGrid-provided P-256 public key. BLOCKER: SendGrid dashboard must enable Signed Event Webhook → POST URL https://api.caseops.ai/api/sendgrid/events with bounce/dropped/spam_report/unsubscribe/group_unsubscribe events. Both per docs/runbooks/sendgrid-event-webhook.md.",
            "L5 — provider integration code-complete while runtime/provider config remained incomplete.",
        ],
        [
            "BUG-039 (Outlook sync-all)",
            "Operational drift / EH-PROV-02",
            "P1",
            "Calendar / Outlook integration",
            "Bounded manual sync vs durable automation",
            "MatterCalendarSyncCard offered single-hearing sync but no bulk sync. Users with 50+ hearings had no path to refresh the entire matter without per-hearing clicks. No durable automation either (Temporal not yet landed).",
            "Added POST /api/matters/{id}/calendar/sync-all endpoint. Bounded: connection check first (409 if no connection), then range check (400 if invalid). Pre-queries existing CalendarEventSync.source_id set so per-row outcomes classify created vs updated. Response includes durable_automation:'blocked_pending_temporal' literal so callers cannot mistake bounded manual sync for continuous automation. Frontend MatterCalendarSyncCard adds 'Sync all hearings' button with disabled state when no connection.",
            "Audited single-hearing sync path (unchanged). Audited Outlook connection lifecycle (unchanged). Confirmed durable_automation literal mirrored in single-sync response.",
            "test_calendar_sync_all.py — connection-required 409, no-hearings 200 with empty results, created vs updated classification, error per row, tenant-isolation cross-matter 404, audit + activity recorded.",
            "MatterCalendarSyncCard.test.tsx — Sync-all button disabled when no connection; enabled state triggers mutation; success toast; error toast.",
            "tests/e2e/hari-2026-05-09-outlook-sync.spec.ts (local: API endpoint mounted, response shape, durable_automation literal). tests/e2e/hari-2026-05-09-outlook-sync-prod.spec.ts (prod: API endpoint mounted on caseops.ai).",
            "Local + prod-Playwright (prod run pending deploy of merged SHA).",
            "Per PR #23 head — see 'Verification Matrix' sheet for SHA.",
            "https://github.com/mishrasanjeev/caseops/pull/23",
            "Partially fixed pending merge + deploy + prod-Playwright run on deployed SHA",
            "Durable automation explicitly deferred to Temporal — the durable_automation:'blocked_pending_temporal' literal is the enterprise-readable source of truth. testMatch overlap on Playwright configs.",
            "L1 — backend/component proof was treated as enough; L5 — durable automation gap is honest in the response.",
        ],
        [
            "Adjacent: linked-record selectors",
            "Audit during sweep (no separate bug ID)",
            "P2",
            "Cross-cutting / Audit",
            "Pattern audit (linked-record selectors with no inline create path)",
            "Same shape as BUG-032 — selector dropdowns query a model without offering an inline create path. Audited matter Notes selector, Tasks selector, MatterAttachment Linked-task selector, Hearings selector. All have create paths reachable from the relevant page (no defects found this sweep), but the audit pattern is now codified in BUG_REOPEN_LEARNINGS_2026-05-09.md L4.",
            "Codified the audit pattern.",
            "Codified the audit pattern.",
            "n/a (no fix this sweep — audit pass).",
            "n/a",
            "n/a",
            "n/a",
            "n/a",
            "n/a",
            "Implemented (audit clean, pattern codified)",
            "Re-audit on every linked-record schema addition.",
            "L4 — codified going forward.",
        ],
        [
            "Adjacent: email-link route-exists guard",
            "Audit during sweep (no separate bug ID)",
            "P2",
            "Cross-cutting / Auth / Audit",
            "Pattern audit (backend-generated URL points at non-existent frontend route)",
            "Same shape as BUG-033. Audited every services/*_mailer.py + services/notifications/*.py reference to settings.web_base_url. Found /account/setup + /account/reset-password (both fixed in PR #21). Other links audited: matter detail, hearing detail, document detail, portal grant accept — all resolve to real routes today.",
            "Codified the audit pattern.",
            "Codified the audit pattern.",
            "n/a (no additional fix this sweep — audit pass).",
            "n/a",
            "n/a",
            "n/a",
            "n/a",
            "n/a",
            "Implemented (audit clean, pattern codified)",
            "Going forward: every new email-link generator needs a paired tests/e2e/* route-exists test.",
            "L2 — codified going forward.",
        ],
    ]

    for r in rows:
        ws.append(r)

    _wrap_rows(ws)
    widths = [22, 28, 10, 32, 38, 40, 60, 40, 40, 32, 40, 30, 22, 50, 38, 50, 38]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_verification_matrix_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Verification Matrix")
    headers = [
        "Bug/Area",
        "PR",
        "Branch",
        "Backend Verification Command",
        "Frontend Verification Command",
        "Local Playwright Spec",
        "Prod Playwright Spec",
        "Pre-merge Local Result",
        "Post-deploy Prod Result",
    ]
    _add_header(ws, headers)
    rows = [
        [
            "BUG-032",
            "#24",
            "fix/bug-032-hearing-order-upload",
            "scripts/verify-backend.sh tests/test_matter_court_orders_create.py",
            "npm --prefix apps/web run test -- page.test",
            "tests/e2e/hari-2026-05-09-bug-032.spec.ts",
            "tests/e2e/hari-2026-05-09-bug-032-prod.spec.ts",
            "Backend 7/7 green; vitest page.test 5/5 green; local Playwright pending re-run after merge-base rebase.",
            "Pending — runs after merge + deploy via Prod verification (Playwright) workflow.",
        ],
        [
            "BUG-033",
            "#21",
            "fix/bug-033-account-setup-links",
            "scripts/verify-backend.sh tests/test_employee_mailer.py tests/test_password_reset.py",
            "npm --prefix apps/web run test -- AccountSetupForm ResetPasswordForm",
            "tests/e2e/hari-2026-05-09-bug-033.spec.ts",
            "tests/e2e/hari-2026-05-09-bug-033-prod.spec.ts",
            "Backend green (auth + mailer); vitest green; local Playwright route-renders green.",
            "Pending — runs after merge + deploy.",
        ],
        [
            "BUG-034",
            "#20",
            "fix/bug-034-protected-capability-catalog",
            "scripts/verify-backend.sh tests/test_capabilities_catalog.py tests/test_custom_roles_create_protected_capability.py",
            "npm --prefix apps/web run test -- RoleCapabilityCheckbox",
            "tests/e2e/hari-2026-05-09-bugs.spec.ts",
            "tests/e2e/hari-2026-05-09-prod.spec.ts",
            "Backend green; vitest green; local Playwright catalog-shape + UI-disabled green.",
            "Pending — runs after merge + deploy.",
        ],
        [
            "BUG-038 (SendGrid)",
            "#22",
            "fix/sendgrid-webhook-delivery-visibility",
            "scripts/verify-backend.sh tests/test_sendgrid_webhook.py tests/test_email_suppression.py tests/test_auth_flow_mailers_bypass_suppression.py",
            "n/a (server-only)",
            "n/a (server-only)",
            "n/a (server-only — operator-side runbook is the proof artifact)",
            "12/12 backend tests green; Alembic upgrade + downgrade dry-run both clean against SQLite.",
            "Pending — Secret Manager value + SendGrid dashboard config + curl probe per docs/runbooks/sendgrid-event-webhook.md after deploy.",
        ],
        [
            "BUG-039 (Outlook)",
            "#23",
            "fix/outlook-sync-all",
            "scripts/verify-backend.sh tests/test_calendar_sync_all.py",
            "npm --prefix apps/web run test -- MatterCalendarSyncCard",
            "tests/e2e/hari-2026-05-09-outlook-sync.spec.ts",
            "tests/e2e/hari-2026-05-09-outlook-sync-prod.spec.ts",
            "Backend green; vitest green; local Playwright endpoint-shape green.",
            "Pending — runs after merge + deploy.",
        ],
    ]
    for r in rows:
        ws.append(r)
    _wrap_rows(ws)
    widths = [22, 8, 42, 55, 50, 45, 45, 55, 55]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_reopen_learnings_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Reopen Learnings")
    headers = ["ID", "Theme", "Pattern", "This-batch instance", "Going-forward rule"]
    _add_header(ws, headers)
    rows = [
        [
            "L1",
            "Backend/component proof was treated as enough",
            "Pytest passed, vitest passed, build succeeded, deploy returned exit 0 — declared the bug fixed. None of those touch the deployed surface a real user sees.",
            "All five PRs in this sweep depend on the prod-Playwright spec to catch a regression after deploy. Only BUG-032 had its full user-visible workflow pre-flighted before PR open.",
            "Every PR carries a *-prod.spec.ts wired into both Playwright configs. Verdict is Partially fixed until that prod spec passes against the deployed commit SHA. No 'Properly fixed' before that proof.",
        ],
        [
            "L2",
            "Email links generated for routes that did not exist",
            "Backend service generates a URL with absolute confidence and ships. Months later the user clicks the link, gets a 404, the bug reopens as 'auth flow has been broken the whole time'.",
            "BUG-033 — /account/setup + /account/reset-password were generated by employee_mailer.py since LW-S5 but the Next.js routes never existed.",
            "Every email-link generator needs a paired tests/e2e/* route-exists test. Reverse linter: grep for web_base_url should be matched by a corresponding route-exists test.",
        ],
        [
            "L3",
            "Backend authorization rules not represented in UI controls",
            "Backend has a hard rule (StrEnum, allow-list, require_capability). UI has a free-text input or unfiltered checkboxes. User submits something the backend will reject, gets a generic 403 toast, files a bug.",
            "BUG-034 — non-delegable capabilities rejected on submit with 403 because the catalog schema only carried owner_only.",
            "Every backend allow-list / non-delegable list / capability partition needs a parallel field on the catalog response so the UI can gate before submit. Disable the control; do not rely on the toast.",
        ],
        [
            "L4",
            "Linked-record metadata existed without a complete create/upload workflow",
            "A schema field exists. A read path renders it. A selector dropdown queries it. There is no create path. The field stays empty for any matter that hasn't had the upstream sync run.",
            "BUG-032 — MatterCourtOrder rows could only come from court-sync; documents-page Linked-order selector permanently empty for matters without sync.",
            "Every linked-record selector needs a discoverable inline create path mounted in the empty state AND in a normal header/toolbar location. Empty state without a create action = defect.",
        ],
        [
            "L5",
            "Provider integration code-complete while runtime/provider config remained incomplete",
            "Code lands, pytest passes against a stubbed provider, PR closes. Infra manifest, Secret Manager, provider dashboard never get done. First real provider event hits the endpoint, verifier returns 503, audit trail is lost for weeks.",
            "BUG-038 — SendGrid webhook code is complete but Secret Manager value + SendGrid dashboard signed-event config remain operator-side action.",
            "A provider integration is not 'ready to ship' until: (a) infra manifest references all required secrets/env declaratively, (b) runbook documents exact operator-side steps, (c) runbook executed at least once, (d) canonical deploy script is the path of record. Anything short is a STRICT_ENTERPRISE_GAP_TASKLIST.md row.",
        ],
        [
            "L6",
            "UX-as-fix substituted for workflow-as-fix",
            "User reports 'I can't complete X.' We respond by softening the error message, adding a banner, redirecting away from the 404, tweaking copy. User still cannot complete X. Bug reopens.",
            "Caught early this sweep — BUG-032 could have shipped as 'empty state explains why orders are missing' without an actual create path. Did not.",
            "A bug is Properly fixed only when the user can complete the intended workflow on the deployed surface. 'Improved copy' / 'added a banner' / 'redirected' alone are Partially fixed at best. Disable the button, do not just toast the error.",
        ],
    ]
    for r in rows:
        ws.append(r)
    _wrap_rows(ws)
    widths = [6, 50, 70, 60, 70]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_pending_prod_proof_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Pending Prod Proof")
    headers = [
        "Bug/Area",
        "PR",
        "Closure step",
        "Owner",
        "Blocker",
        "Acceptance criterion",
    ]
    _add_header(ws, headers)
    rows = [
        ["BUG-032", "#24", "Merge to main", "Reviewer", "Awaiting review", "Squash-merge with the PR-body verification matrix preserved."],
        ["BUG-032", "#24", "Deploy via scripts/deploy-prod.sh", "Operator", "Post-merge", "Deploy script exit 0; revision tag matches merged SHA."],
        ["BUG-032", "#24", "Run hari-2026-05-09-bug-032-prod.spec.ts on deployed SHA", "Prod verification (Playwright) workflow", "Post-deploy", "Both tests green against caseops.ai with disposable QA-tenant artefacts."],
        ["BUG-033", "#21", "Merge to main", "Reviewer", "Awaiting review", "Squash-merge."],
        ["BUG-033", "#21", "Deploy via scripts/deploy-prod.sh", "Operator", "Post-merge", "Deploy script exit 0."],
        ["BUG-033", "#21", "Run hari-2026-05-09-bug-033-prod.spec.ts on deployed SHA", "Prod-Playwright workflow", "Post-deploy", "Route-renders test green; happy-path with debug_token requires DEBUG_TOKEN_ENVS broadening (separate task)."],
        ["BUG-034", "#20", "Merge to main", "Reviewer", "Awaiting review", "Squash-merge."],
        ["BUG-034", "#20", "Deploy via scripts/deploy-prod.sh", "Operator", "Post-merge", "Deploy script exit 0."],
        ["BUG-034", "#20", "Run hari-2026-05-09-prod.spec.ts on deployed SHA", "Prod-Playwright workflow", "Post-deploy", "Catalog endpoint reports custom_role_delegable + protected_reason on prod; UI disables protected caps before submit."],
        ["BUG-038 (SendGrid)", "#22", "Merge to main", "Reviewer", "Awaiting review", "Squash-merge with EmailSuppression migration."],
        ["BUG-038 (SendGrid)", "#22", "Deploy via scripts/deploy-prod.sh", "Operator", "Post-merge", "Deploy script exit 0; Alembic migration applied; manifest references caseops-sendgrid-webhook-public-key."],
        ["BUG-038 (SendGrid)", "#22", "Create Secret Manager value caseops-sendgrid-webhook-public-key", "Operator", "Awaiting SendGrid-provided P-256 public key", "Secret value present and Cloud Run revision rolls forward."],
        ["BUG-038 (SendGrid)", "#22", "Configure SendGrid dashboard Signed Event Webhook", "Operator", "Awaiting Secret Manager value", "Settings → Mail Settings → Event Webhook → POST URL https://api.caseops.ai/api/sendgrid/events; bounce/dropped/spam_report/unsubscribe/group_unsubscribe enabled."],
        ["BUG-038 (SendGrid)", "#22", "Run end-to-end webhook curl probe per runbook", "Operator", "Awaiting dashboard config", "200 on signed valid event; 503 on missing/invalid signature; suppression row appears in tenant-scoped table."],
        ["BUG-039 (Outlook)", "#23", "Merge to main", "Reviewer", "Awaiting review", "Squash-merge."],
        ["BUG-039 (Outlook)", "#23", "Deploy via scripts/deploy-prod.sh", "Operator", "Post-merge", "Deploy script exit 0."],
        ["BUG-039 (Outlook)", "#23", "Run hari-2026-05-09-outlook-sync-prod.spec.ts on deployed SHA", "Prod-Playwright workflow", "Post-deploy", "Endpoint mounted on api.caseops.ai; response carries durable_automation:'blocked_pending_temporal' literal."],
        ["Cross-PR", "#20/#21/#23/#24", "Resolve testMatch overlap on second/third/fourth PR rebase", "Whoever rebases", "First merge happens", "All four hari-2026-05-09-*.spec.ts entries present in playwright.app.config.ts AND playwright.prod-ram.config.ts after each rebase."],
    ]
    for r in rows:
        ws.append(r)
    _wrap_rows(ws)
    widths = [24, 16, 50, 32, 32, 70]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def main() -> None:
    wb = Workbook()
    build_summary_sheet(wb)
    build_bug_details_sheet(wb)
    build_verification_matrix_sheet(wb)
    build_reopen_learnings_sheet(wb)
    build_pending_prod_proof_sheet(wb)

    out_dir = os.path.dirname(OUT_PATH)
    os.makedirs(out_dir, exist_ok=True)
    wb.save(OUT_PATH)
    print(f"Wrote {OUT_PATH}")


if __name__ == "__main__":
    main()
