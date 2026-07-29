from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[1]
SOURCE = Path(r"C:\Users\mishr\Downloads\CaseOps_BugsRam29Jul2026.xlsx")
OUTPUT = ROOT / "outputs" / "CaseOps_Bug_Fix_Summary_Ram29Jul2026.xlsx"

NAVY = "17324D"
TEAL = "0F766E"
PALE_TEAL = "D9F0EE"
PALE_GREEN = "DCFCE7"
PALE_AMBER = "FFF4D6"
PALE_RED = "FEE2E2"
WHITE = "FFFFFF"
INK = "243447"
MUTED = "5B6B7A"
GRID = "D6DEE6"


def title(ws, text: str, subtitle: str, end_col: int) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)
    ws.cell(1, 1).value = text
    ws.cell(1, 1).font = Font(name="Aptos Display", size=18, bold=True, color=WHITE)
    ws.cell(1, 1).fill = PatternFill("solid", fgColor=NAVY)
    ws.cell(1, 1).alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 30
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=end_col)
    ws.cell(2, 1).value = subtitle
    ws.cell(2, 1).font = Font(name="Aptos", size=10, color=MUTED, italic=True)
    ws.cell(2, 1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[2].height = 36


def header(ws, row: int, start_col: int, end_col: int) -> None:
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row, col)
        cell.font = Font(name="Aptos", size=10, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=TEAL)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = Border(bottom=Side(style="medium", color=NAVY))
    ws.row_dimensions[row].height = 36


def body(ws, first_row: int, last_row: int, first_col: int, last_col: int, height: int = 78) -> None:
    for row in ws.iter_rows(
        min_row=first_row,
        max_row=last_row,
        min_col=first_col,
        max_col=last_col,
    ):
        for cell in row:
            cell.font = Font(name="Aptos", size=10, color=INK)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = Border(bottom=Side(style="hair", color=GRID))
    for row in range(first_row, last_row + 1):
        ws.row_dimensions[row].height = height


def add_table(ws, name: str, ref: str) -> None:
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def build() -> Path:
    source = load_workbook(SOURCE, data_only=False).active
    source_rows = [
        list(row)
        for row in source.iter_rows(min_row=2, max_col=16, values_only=True)
        if row[0]
    ]
    if len(source_rows) != 2:
        raise RuntimeError(f"Expected exactly 2 non-empty workbook rows, found {len(source_rows)}")

    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"
    assessment = wb.create_sheet("Item Assessment")
    evidence = wb.create_sheet("Test Evidence")
    reopen = wb.create_sheet("Why Cases Reopen")
    regression = wb.create_sheet("Regression Matrix")
    source_sheet = wb.create_sheet("Source Items")
    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
        ws.sheet_view.zoomScale = 90

    title(
        summary,
        "CaseOps Bug Fix Summary — Ram 29 Jul 2026",
        "Two workbook rows were assessed against the CaseOps repository and live legal tenant. Production credentials are intentionally omitted. Fixed means the exact dated Playwright proof passed after the deployed image/revision was verified.",
        10,
    )
    cards = [
        ("A4", "Workbook rows", "A5", '=COUNTA(\'Item Assessment\'!$A$7:$A$8)'),
        ("C4", "Valid bugs", "C5", '=COUNTIF(\'Item Assessment\'!$E$7:$E$8,"Valid bug")'),
        ("E4", "Fixed / revalidated", "E5", '=COUNTIF(\'Item Assessment\'!$F$7:$F$8,"Fixed / revalidated")'),
        ("G4", "Prod Playwright pass", "G5", '=COUNTIF(\'Test Evidence\'!$D$7:$D$14,"PASS*")'),
        ("I4", "Lifecycle result", "I5", '=COUNTIF(\'Regression Matrix\'!$F$7:$F$11,"PASS")'),
    ]
    for label_cell, label, value_cell, formula in cards:
        summary[label_cell] = label
        summary[label_cell].font = Font(name="Aptos", size=10, bold=True, color=MUTED)
        summary[value_cell] = formula
        summary[value_cell].font = Font(name="Aptos Display", size=18, bold=True, color=NAVY)
        summary[value_cell].fill = PatternFill("solid", fgColor=PALE_TEAL)
        summary[value_cell].alignment = Alignment(horizontal="center", vertical="center")
    summary.row_dimensions[5].height = 34
    summary.append([])
    summary.append(["Area", "Finding", "Verdict", "Evidence / next action"])
    header(summary, 7, 1, 4)
    summary_rows = [
        ["BUG-001", "Judge Aliases was missing from the shared Admin navigation registry; the page and landing-page action already existed.", "Fixed / revalidated", "Valid bug. Existing shared Sidebar fix retained; live desktop + mobile Playwright passed after fd1238c deployment."],
        ["BUG-002", "The Admin header nested eight actions inside a non-wrapping flex row; Provider ops extended 3.484px beyond a 360px viewport.", "Fixed / revalidated", "Valid bug. Admin action group and shared PageHeader slot now shrink/wrap; live 360px Playwright passed."],
        ["Case reopening", "Disposed -> Intake is an intentional controlled transition; unexpected reactivation was not reproduced after deployment.", "Lifecycle revalidated", "Live lifecycle Playwright passed. Generic PATCH, stale writes, child resurrection, and operational-view leakage remained blocked."],
        ["Release discipline", "The source tree and the running artifact had previously drifted, allowing a stale build to look like a reopened bug.", "Permanent control added", "Exact image tag/revision, migration gate, and same dated production spec are now required evidence."],
    ]
    for row in summary_rows:
        summary.append(row)
    body(summary, 8, 11, 1, 4)
    for col, width in {"A": 22, "B": 68, "C": 28, "D": 72}.items():
        summary.column_dimensions[col].width = width
    summary.auto_filter.ref = "A7:D11"
    summary.freeze_panes = "A7"

    headers = ["ID", "Reported type", "Module", "Environment", "Classification", "Verdict", "Reported summary", "Root cause", "Fix / disposition", "Regression proof", "Production identity", "Notes"]
    title(assessment, "Item-by-Item Assessment", "Both source rows are valid CaseOps bugs. BUG-001 was a prior fix revalidation; BUG-002 required the responsive code change in fd1238c.", len(headers))
    for _ in range(3):
        assessment.append([])
    assessment.append(headers)
    header(assessment, 6, 1, len(headers))
    rows = [
        [
            "BUG-001", "UI / Responsive", "Admin & Governance / Judge Aliases", "Chrome desktop + Android-sized mobile", "Valid bug", "Fixed / revalidated",
            str(source_rows[0][6]).replace("\n", " "),
            "The page, API, and Admin landing action existed, but the shared Sidebar NAV registry omitted /app/admin/judge-aliases. A stale local .next artifact also made the already-fixed source look broken during retest.",
            "Retained the capability-gated shared Sidebar item. No duplicate page-specific navigation was added. Rebuilt the production web artifact and revalidated both navigation surfaces.",
            "Local Playwright 2/2; production Playwright BUG-001 PASS after fd1238c. Existing Sidebar unit coverage also passes.",
            "caseops-web-00194-5qt; image fd1238c; 100% traffic",
            "Valid bug. The code fix predates this commit; this run closes the revalidation gap.",
        ],
        [
            "BUG-002", "Responsive UI", "Admin & Governance / Admin landing page", "Chrome Android-sized 360px viewport", "Valid bug", "Fixed / revalidated",
            str(source_rows[1][6]).replace("\n", " "),
            "AdminPage passed a single inner flex row as PageHeader actions. That inner row had no flex-wrap/min-width constraints, so its intrinsic width pushed Provider ops and later actions outside the viewport.",
            "Added min-w-0/full-width/flex-wrap responsive constraints to the shared PageHeader action slot and Admin action group. This fixes the shared layout boundary rather than hiding individual links.",
            "Pre-deploy production reproduced Provider ops right edge at 363.484px. Local production-build Playwright 2/2 and post-deploy production Playwright 2/2 passed.",
            "caseops-web-00194-5qt; image fd1238c; 100% traffic",
            "No API change; canonical migration/deploy gate still passed.",
        ],
    ]
    for row in rows:
        assessment.append(row)
    body(assessment, 7, 8, 1, len(headers), height=104)
    for index, width in enumerate([16, 20, 34, 32, 20, 22, 54, 68, 68, 64, 44, 40], start=1):
        assessment.column_dimensions[get_column_letter(index)].width = width
    assessment.freeze_panes = "A7"
    assessment.auto_filter.ref = "A6:L8"
    add_table(assessment, "CaseOpsRam29Items", "A6:L8")
    assessment.conditional_formatting.add("F7:F8", FormulaRule(formula=['$F7="Fixed / revalidated"'], fill=PatternFill("solid", fgColor=PALE_GREEN)))

    evidence_headers = ["Date", "Surface", "Exact proof artifact", "Result", "What it proves", "Limitations / caveat"]
    title(evidence, "Verification Evidence", "Evidence is separated by surface. Production proof includes the deployed identity and the exact Playwright spec; passwords are not recorded.", len(evidence_headers))
    for _ in range(3):
        evidence.append([])
    evidence.append(evidence_headers)
    header(evidence, 6, 1, len(evidence_headers))
    evidence_rows = [
        ["2026-07-29", "Focused web unit", "apps/web/components/app/Sidebar.test.tsx + app/app/admin/page.test.tsx", "PASS: 12/12", "Capability filtering, Admin page actions, and existing UI contracts pass.", "Unit tests are not deployment proof."],
        ["2026-07-29", "Web typecheck/build", "npm run typecheck:web; npm run build:web", "PASS", "Current source typechecks and produces a fresh Next production build.", "Local artifact must still be deployed and rechecked."],
        ["2026-07-29", "API lifecycle", "apps/api/tests/test_matter_lifecycle.py", "PASS: 16/16", "Terminal-state, CAS, lifecycle endpoint, child neutralization, and legacy-row controls pass.", "SQLite/API test evidence is not live tenant evidence."],
        ["2026-07-29", "Local Playwright", "tests/e2e/ram-2026-07-29-bugs.spec.ts", "PASS: 2/2", "Judge Aliases desktop/mobile discoverability and all Admin action bounds at 360px pass on a fresh local legal tenant.", "Local proof is not production proof."],
        ["2026-07-29", "Production pre-deploy Playwright", "tests/e2e/ram-2026-07-29-prod.spec.ts", "PASS: 1/2; BUG-002 FAIL", "BUG-001 passed; BUG-002 reproduced live with Provider ops at 363.484px right edge before deployment.", "This was the reproduction run, not a release sign-off."],
        ["2026-07-29", "Production deployment", "fd1238c; caseops-api-00214-45z; caseops-web-00194-5qt", "PASS", "Cloud Build, migration job, recurring job repinning, 100% traffic, staleness sweep, health, and clamAV sidecar guard passed.", "Health is availability evidence; image/revision is the release identity."],
        ["2026-07-29", "Production Playwright", "tests/e2e/ram-2026-07-29-prod.spec.ts", "PASS: 2/2", "Supplied tester account passed both workbook regressions after fd1238c reached 100% traffic.", "Production credentials omitted from this workbook."],
        ["2026-07-29", "Production lifecycle Playwright", "tests/e2e/ram-2026-07-15-prod.spec.ts", "PASS: 3/3", "Active creation, notice flow, and full dispose/reopen lifecycle passed; stale writes and resurrection stayed blocked.", "Created test records were returned to terminal state by the suite's cleanup."],
    ]
    for row in evidence_rows:
        evidence.append(row)
    body(evidence, 7, 14, 1, len(evidence_headers), height=84)
    for index, width in enumerate([16, 28, 66, 24, 70, 58], start=1):
        evidence.column_dimensions[get_column_letter(index)].width = width
    evidence.freeze_panes = "A7"
    evidence.auto_filter.ref = "A6:F14"
    add_table(evidence, "CaseOpsRam29Evidence", "A6:F14")
    evidence.conditional_formatting.add("D7:D14", FormulaRule(formula=['LEFT($D7,4)="PASS"'], fill=PatternFill("solid", fgColor=PALE_GREEN)))
    evidence.conditional_formatting.add("D7:D14", FormulaRule(formula=['ISNUMBER(SEARCH("FAIL",$D7))'], fill=PatternFill("solid", fgColor=PALE_RED)))

    reopen_headers = ["Control", "Finding", "Permanent rule", "Evidence"]
    title(reopen, "Why Cases Reopen", "Brutal analysis: distinguish intentional lifecycle reopening from release drift and from child/status resurrection.", len(reopen_headers))
    for _ in range(3):
        reopen.append([])
    reopen.append(reopen_headers)
    header(reopen, 6, 1, len(reopen_headers))
    reopen_rows = [
        ["Intentional transition", "Disposed -> Intake is a controlled product action, not an accidental reopen.", "Only the dedicated lifecycle endpoint may perform it; it requires expected status, updated_at, reason, and archive capability.", "Production lifecycle spec PASS."],
        ["Generic metadata writes", "A stale browser or generic PATCH must not replay Active after disposal.", "Lock the parent and reject inactive/terminal metadata writes plus stale CAS tokens with HTTP 409.", "Production lifecycle stale-write and generic reactivation assertions PASS."],
        ["Child resurrection", "Tasks, deadlines, hearings, reminders, and provider rows can make a disposed case look operational even when the parent is terminal.", "Dispose/reopen neutralizes open children, sets durable external tombstones, and blocks resurrection patches.", "Production lifecycle child suppression/resurrection assertions PASS."],
        ["Release drift", "A source fix can appear reopened when a stale .next or Cloud Run image is serving old code.", "Build current source, verify exact image/revision, and run the same dated Playwright spec after deployment.", "Local stale-build failure + Cloud Run fd1238c staleness sweep PASS."],
        ["Responsive drift", "A nested flex row can hide later controls only at mobile widths while desktop appears healthy.", "Mobile regressions assert every grouped action's rendered bounds, not only route existence.", "Pre-deploy production reproduced 363.484px; post-deploy 360px proof PASS."],
        ["Process failure", "Previous shallow fixes trusted a green local or partial proof and did not bind verdicts to a release artifact.", "Fail closed: no 'fixed' label without code, test, deployment identity, production replay, and cleanup evidence.", "Permanent rules added to AGENTS.md and docs/BUG_REOPEN_LEARNINGS_2026-07-29_RAM.md."],
    ]
    for row in reopen_rows:
        reopen.append(row)
    body(reopen, 7, 12, 1, len(reopen_headers), height=92)
    for index, width in enumerate([28, 68, 72, 58], start=1):
        reopen.column_dimensions[get_column_letter(index)].width = width
    reopen.freeze_panes = "A7"
    reopen.auto_filter.ref = "A6:D12"
    add_table(reopen, "CaseOpsRam29ReopenControls", "A6:D12")

    regression_headers = ["Regression area", "Exact suite", "Scenario", "Expected invariant", "Environment", "Result"]
    title(regression, "Regression Matrix", "These are the durable checks that prevent the same classes of bug from returning elsewhere in the product.", len(regression_headers))
    for _ in range(3):
        regression.append([])
    regression.append(regression_headers)
    header(regression, 6, 1, len(regression_headers))
    regression_rows = [
        ["Admin navigation", "tests/e2e/ram-2026-07-29-bugs.spec.ts / prod.spec.ts", "Desktop + mobile Judge aliases navigation and destination render", "Shared capability-gated route is visible, clickable, and destination heading renders", "Local + production", "PASS"],
        ["Admin responsive actions", "tests/e2e/ram-2026-07-29-bugs.spec.ts / prod.spec.ts", "All eight Admin actions at 360px", "Every action is visible, has the expected href, and stays within viewport bounds", "Local + production", "PASS"],
        ["Terminal lifecycle", "tests/e2e/ram-2026-07-15-prod.spec.ts", "Dispose, stale edit, generic reactivation, child writes", "409 on unsafe writes; disposed stays out of Today/Calendar", "Production", "PASS"],
        ["Controlled reopen", "tests/e2e/ram-2026-07-15-prod.spec.ts", "Disposed -> Intake -> Active with reload", "Only dedicated lifecycle action reopens; cancelled children do not revive; Active persists", "Production", "PASS"],
        ["Backend lifecycle suite", "apps/api/tests/test_matter_lifecycle.py", "State matrix, DB invariant, legacy rows, conflict history", "16 tests pass; status/is_active/lifecycle_version remain coherent", "Local API", "PASS"],
    ]
    for row in regression_rows:
        regression.append(row)
    body(regression, 7, 11, 1, len(regression_headers), height=82)
    for index, width in enumerate([28, 52, 62, 70, 24, 16], start=1):
        regression.column_dimensions[get_column_letter(index)].width = width
    regression.freeze_panes = "A7"
    regression.auto_filter.ref = "A6:F11"
    add_table(regression, "CaseOpsRam29RegressionMatrix", "A6:F11")
    regression.conditional_formatting.add("F7:F11", FormulaRule(formula=['$F7="PASS"'], fill=PatternFill("solid", fgColor=PALE_GREEN)))

    source_headers = [
        "Bug ID", "Severity", "Type", "Module", "Sub Module", "Page URL", "Summary",
        "Steps to Reproduce", "Expected Result", "Actual Result", "Status", "Environment",
        "Reported By", "Reported On", "Screenshot",
    ]
    title(source_sheet, "Source Workbook Rows", "Copied from CaseOps_BugsRam29Jul2026.xlsx with the Credentials column intentionally excluded.", len(source_headers))
    for _ in range(3):
        source_sheet.append([])
    source_sheet.append(source_headers)
    header(source_sheet, 6, 1, len(source_headers))
    for row in source_rows:
        safe = row[:14] + [row[15]]
        source_sheet.append(safe)
    body(source_sheet, 7, 8, 1, len(source_headers), height=96)
    for index, width in enumerate([14, 12, 22, 26, 24, 46, 58, 58, 58, 58, 14, 28, 16, 16, 24], start=1):
        source_sheet.column_dimensions[get_column_letter(index)].width = width
    source_sheet.freeze_panes = "A7"
    source_sheet.auto_filter.ref = "A6:O8"
    add_table(source_sheet, "CaseOpsRam29SourceItems", "A6:O8")

    for ws in wb.worksheets:
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.page_margins.left = 0.25
        ws.page_margins.right = 0.25
        ws.page_margins.top = 0.5
        ws.page_margins.bottom = 0.5
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    return OUTPUT


if __name__ == "__main__":
    print(build())
