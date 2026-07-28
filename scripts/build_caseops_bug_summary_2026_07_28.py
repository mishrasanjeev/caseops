from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[1]
SOURCE = Path(r"C:\Users\mishr\Downloads\CaseOps_Bugs_Ram28Jul2026.xlsx")
OUTPUT = ROOT / "outputs" / "CaseOps_Bug_Fix_Summary_Ram28Jul2026.xlsx"

NAVY = "17324D"
TEAL = "0F766E"
PALE_TEAL = "D9F0EE"
PALE_GREEN = "DCFCE7"
PALE_AMBER = "FFF4D6"
WHITE = "FFFFFF"
INK = "243447"
MUTED = "5B6B7A"
GRID = "D6DEE6"


def title(ws, text: str, subtitle: str, end_col: int) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)
    ws.cell(1, 1).value = text
    ws.cell(1, 1).font = Font(name="Aptos Display", size=18, bold=True, color=WHITE)
    ws.cell(1, 1).fill = PatternFill("solid", fgColor=NAVY)
    ws.cell(1, 1).alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 30
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=end_col)
    ws.cell(2, 1).value = subtitle
    ws.cell(2, 1).font = Font(name="Aptos", size=10, color=MUTED, italic=True)
    ws.cell(2, 1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[2].height = 32


def header(ws, row: int, end_col: int) -> None:
    for col in range(1, end_col + 1):
        cell = ws.cell(row, col)
        cell.font = Font(name="Aptos", size=10, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=TEAL)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = Border(bottom=Side(style="medium", color=NAVY))
    ws.row_dimensions[row].height = 34


def body(ws, first_row: int, last_row: int, end_col: int) -> None:
    for row in ws.iter_rows(min_row=first_row, max_row=last_row, min_col=1, max_col=end_col):
        for cell in row:
            cell.font = Font(name="Aptos", size=10, color=INK)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = Border(bottom=Side(style="hair", color=GRID))
    for row in range(first_row, last_row + 1):
        ws.row_dimensions[row].height = 74


def build() -> Path:
    source = load_workbook(SOURCE, data_only=False).active
    source_values = next(source.iter_rows(min_row=2, max_row=2, values_only=True))
    source_id = str(source_values[0])
    source_summary = str(source_values[6]).replace("\n", " ")
    source_steps = str(source_values[7]).replace("\n", " ")
    source_expected = str(source_values[8]).replace("\n", " ")
    source_actual = str(source_values[9]).replace("\n", " ")

    rows = [
        [
            source_id,
            "Workbook row",
            str(source_values[2]),
            str(source_values[3]),
            "Valid CaseOps bug",
            "Properly fixed",
            source_summary,
            "Shared Sidebar NAV registry omitted /app/admin/judge-aliases. The route, API, Admin landing-page action, and page tests already existed.",
            "Added a capability-gated Judge aliases item to the shared Sidebar registry. Because SidebarBody is reused by the mobile drawer, both surfaces now stay aligned.",
            "Desktop + mobile navigation and real destination page are covered by tests/e2e/ram-2026-07-28-bugs.spec.ts and tests/e2e/ram-2026-07-28-prod.spec.ts.",
            "Candidate web image 7495bc6; Cloud Run revision caseops-web-00191-vn9 at 100% traffic.",
            "Production proof passed with supplied tester account; password intentionally omitted.",
        ],
        [
            "DEPLOY-GATE-2026-07-28",
            "Adjacent release finding",
            "Production migration job cannot locate database revision 20260723_0001.",
            "Production deployment pipeline",
            "Valid release blocker",
            "Inconclusive",
            "Canonical full deploy was blocked before API rollout because the production database points to an Alembic revision absent from the repository image.",
            "gcloud run jobs execute caseops-migrate-job failed with exit code 255: Can't locate revision identified by '20260723_0001'.",
            "Did not bypass the migration gate. Deployed only the web image because BUG-001 is web-only; API traffic was not changed.",
            "Cloud Run job execution caseops-migrate-job-7mmw2; logs recorded in the deployment audit.",
            "API image built but not deployed; web-only deployment succeeded.",
            "Requires migration-lineage repair before the next full API release.",
        ],
    ]

    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"
    analysis = wb.create_sheet("Analysis")
    evidence = wb.create_sheet("Test Evidence")
    reopen = wb.create_sheet("Reopen Audit")
    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False

    title(summary, "CaseOps Bug Fix Summary — Ram 28 Jul 2026", "One workbook row was reviewed against CaseOps. Verdicts are fail-closed and production credentials/passwords are not recorded.", 10)
    cards = [("A4", "Workbook rows"), ("C4", "Valid CaseOps bugs"), ("E4", "Properly fixed"), ("G4", "Release blockers"), ("I4", "Inconclusive")]
    for cell, label in cards:
        summary[cell] = label
        summary[cell].font = Font(name="Aptos", size=10, bold=True, color=MUTED)
    summary["A5"] = "=COUNTIF('Analysis'!$B$7:$B$8,\"Workbook row\")"
    summary["C5"] = "=COUNTIF('Analysis'!$E$7:$E$8,\"Valid CaseOps bug\")"
    summary["E5"] = "=COUNTIF('Analysis'!$F$7:$F$8,\"Properly fixed\")"
    summary["G5"] = "=COUNTIF('Analysis'!$E$7:$E$8,\"Valid release blocker\")"
    summary["I5"] = "=COUNTIF('Analysis'!$F$7:$F$8,\"Inconclusive\")"
    for cell in [summary["A5"], summary["C5"], summary["E5"], summary["G5"], summary["I5"]]:
        cell.font = Font(name="Aptos Display", size=18, bold=True, color=NAVY)
        cell.fill = PatternFill("solid", fgColor=PALE_TEAL)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    summary.row_dimensions[5].height = 34
    summary.append([])
    summary.append(["Area", "Finding", "Decision", "Evidence / next action"])
    header(summary, 7, 4)
    summary_rows = [
        ["BUG-001 scope", "Judge Aliases was implemented but missing from the shared Admin navigation registry.", "Valid bug; properly fixed", "Web candidate 7495bc6 deployed to 100% and production Playwright passed."],
        ["Regression breadth", "Desktop sidebar and shared mobile drawer now expose the same capability-gated route.", "Covered", "Local and production dated specs assert navigation plus destination page."],
        ["Full deployment", "The API migration job references missing revision 20260723_0001.", "Release blocker", "Repair migration lineage before the next API deployment; do not bypass the gate."],
    ]
    for row in summary_rows:
        summary.append(row)
    body(summary, 8, 10, 4)
    for col, width in {"A": 24, "B": 62, "C": 28, "D": 70}.items():
        summary.column_dimensions[col].width = width
    summary.auto_filter.ref = "A7:D10"
    summary.freeze_panes = "A7"

    analysis_headers = ["ID", "Record Type", "Module", "Reported Environment", "Classification", "Verdict", "Reported Summary", "Root Cause", "Fix / Decision", "Regression Proof", "Deployment Evidence", "Notes"]
    title(analysis, "Item-by-Item Assessment", "The workbook row is separated from an adjacent release blocker discovered while deploying and revalidating it.", len(analysis_headers))
    for _ in range(3):
        analysis.append([])
    analysis.append(analysis_headers)
    header(analysis, 6, len(analysis_headers))
    for row in rows:
        analysis.append(row)
    body(analysis, 7, 8, len(analysis_headers))
    for index, width in enumerate([24, 22, 28, 26, 24, 20, 56, 62, 62, 62, 48, 42], start=1):
        analysis.column_dimensions[get_column_letter(index)].width = width
    analysis.freeze_panes = "A7"
    analysis.auto_filter.ref = "A6:L8"
    table = Table(displayName="CaseOpsAuditItems", ref="A6:L8")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    analysis.add_table(table)
    analysis.conditional_formatting.add("F7:F8", FormulaRule(formula=['$F7="Properly fixed"'], fill=PatternFill("solid", fgColor=PALE_GREEN)))
    analysis.conditional_formatting.add("F7:F8", FormulaRule(formula=['$F7="Inconclusive"'], fill=PatternFill("solid", fgColor=PALE_AMBER)))

    evidence_headers = ["Date", "Surface", "Exact proof artifact", "Result", "What it proves", "Limitations / caveat"]
    title(evidence, "Verification Evidence", "Every fixed verdict is paired with a repeatable user-visible proof artifact and deployed build identity.", len(evidence_headers))
    for _ in range(3):
        evidence.append([])
    evidence.append(evidence_headers)
    header(evidence, 6, len(evidence_headers))
    evidence_rows = [
        ["2026-07-28", "Focused web unit", "apps/web/components/app/Sidebar.test.tsx", "PASS: 5/5", "Capability-gated navigation includes Judge aliases for an allowed custom-role capability set.", "Component test is not deployment proof."],
        ["2026-07-28", "Local Playwright", "tests/e2e/ram-2026-07-28-bugs.spec.ts", "PASS: 1/1", "Fresh local legal tenant: desktop sidebar and 360px mobile drawer both navigate to and render Judge Aliases.", "Local build is not production proof."],
        ["2026-07-28", "Production pre-deploy Playwright", "tests/e2e/ram-2026-07-28-prod.spec.ts", "FAIL as expected", "Reproduced the original defect: signed-in tester could not find the Judge aliases link before deployment.", "Captured against the old production web revision."],
        ["2026-07-28", "Cloud Build", "candidate commit 7495bc6; web image digest sha256:d1832f...", "PASS", "Production web artifact built with TypeScript/Next production build and pushed to Artifact Registry.", "API image was also built but not deployed because migration gate failed."],
        ["2026-07-28", "Production web deployment", "caseops-web-00191-vn9; image tag 7495bc6", "PASS: 100% traffic", "The web-only fix is the live production revision.", "Full canonical API rollout remains blocked by missing Alembic revision 20260723_0001."],
        ["2026-07-28", "Production Playwright", "tests/e2e/ram-2026-07-28-prod.spec.ts", "PASS: 1/1", "Supplied tester account: desktop sidebar, mobile drawer, destination URL, and Judge Aliases heading all passed on live caseops.ai.", "Password omitted; API health proves availability only. Cloud Run image tag proves web candidate identity."],
        ["2026-07-28", "Migration gate", "caseops-migrate-job-7mmw2 logs", "BLOCKED: exit 255", "Prevented an unsafe API rollout; logs report missing revision 20260723_0001.", "Requires a separate migration-lineage repair before a full deploy."],
    ]
    for row in evidence_rows:
        evidence.append(row)
    body(evidence, 7, 13, len(evidence_headers))
    for index, width in enumerate([16, 28, 62, 24, 68, 58], start=1):
        evidence.column_dimensions[get_column_letter(index)].width = width
    evidence.auto_filter.ref = "A6:F13"
    evidence.freeze_panes = "A7"

    reopen_headers = ["Control", "Finding", "Permanent rule"]
    title(reopen, "Why Cases Reopen / Release Drift Audit", "This batch did not introduce a new lifecycle bug, but the deployment investigation exposed why local fixes can appear to reopen in production.", len(reopen_headers))
    for _ in range(3):
        reopen.append([])
    reopen.append(reopen_headers)
    header(reopen, 6, len(reopen_headers))
    reopen_rows = [
        ["User-visible scope", "A route can exist and be linked from an Admin landing page while still being absent from the primary navigation registry.", "Acceptance tests must start at the actual main navigation and cover desktop plus mobile when they share a primitive."],
        ["Root cause of BUG-001", "Sidebar NAV omitted /app/admin/judge-aliases; the page/API were already present.", "Maintain one shared navigation registry and a regression that clicks the link and renders the destination."],
        ["Why fixes appear to reopen", "Previous evidence can prove a local candidate while production continues serving an older or drifted build.", "Pair every fixed verdict with the exact deployed image/revision and the same dated production Playwright spec."],
        ["Migration lineage", "The full deploy was correctly stopped because production alembic_version points to absent revision 20260723_0001.", "Never bypass the migration job; add release checks that the image contains every revision referenced by production before API rollout."],
        ["Fail-closed verdict", "The web-only BUG-001 fix passed production, but the overall API/web release was not certified because the migration gate failed.", "Separate item-level proof from release-level GO/NO-GO and record blockers in the summary."],
    ]
    for row in reopen_rows:
        reopen.append(row)
    body(reopen, 7, 11, len(reopen_headers))
    reopen.column_dimensions["A"].width = 28
    reopen.column_dimensions["B"].width = 72
    reopen.column_dimensions["C"].width = 76
    reopen.auto_filter.ref = "A6:C11"
    reopen.freeze_panes = "A7"

    for ws in wb.worksheets:
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_view.zoomScale = 90
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    return OUTPUT


if __name__ == "__main__":
    print(build())
