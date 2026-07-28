from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[1]
SOURCE = Path(r"C:\Users\mishr\Downloads\Enhancements_Ram28Jul2026.xlsx")
OUTPUT = ROOT / "outputs" / "Bug_Fix_Summary_Ram28Jul2026.xlsx"

NAVY = "17324D"
TEAL = "0F766E"
PALE_TEAL = "D9F0EE"
PALE_AMBER = "FFF4D6"
WHITE = "FFFFFF"
INK = "243447"
MUTED = "5B6B7A"
GRID = "D6DEE6"


def style_title(ws, title: str, subtitle: str, end_col: int) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)
    ws.cell(1, 1).value = title
    ws.cell(1, 1).font = Font(name="Aptos Display", size=18, bold=True, color=WHITE)
    ws.cell(1, 1).fill = PatternFill("solid", fgColor=NAVY)
    ws.cell(1, 1).alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 30
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=end_col)
    ws.cell(2, 1).value = subtitle
    ws.cell(2, 1).font = Font(name="Aptos", size=10, color=MUTED, italic=True)
    ws.cell(2, 1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[2].height = 30


def style_header_row(ws, row: int, start_col: int, end_col: int) -> None:
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row, col)
        cell.font = Font(name="Aptos", size=10, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=TEAL)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = Border(bottom=Side(style="medium", color=NAVY))
    ws.row_dimensions[row].height = 34


def style_body(ws, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    for row in ws.iter_rows(
        min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col
    ):
        for cell in row:
            cell.font = Font(name="Aptos", size=10, color=INK)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = Border(bottom=Side(style="hair", color=GRID))
    for row in range(min_row, max_row + 1):
        ws.row_dimensions[row].height = 66


def build() -> Path:
    source_wb = load_workbook(SOURCE, data_only=False)
    source_rows = list(source_wb.active.iter_rows(min_row=2, max_col=11, values_only=True))

    rows = [
        {
            "id": "1", "reported_type": "Enhancement",
            "summary": str(source_rows[0][1]).replace("\n", " "),
            "module": str(source_rows[0][2]).replace("\n", " "),
            "validity": "Not a CaseOps item",
            "repo_match": "No matching Edumatica or Physical Library module exists in this repository.",
            "verdict": "Inconclusive",
            "root_cause": "External Edumatica scope. The supplied URL and domain model are outside CaseOps.",
            "action": "No unrelated code change. Request the Edumatica repository or deployment context for implementation.",
            "proof": "Repository search: no Physical Library, Student Due Management, Degree Course, or Degree Program surfaces.",
            "production": "Not tested: external product",
            "notes": "Valid-looking enhancement for another product; not a CaseOps enhancement.",
        },
        {
            "id": "2", "reported_type": "Bug",
            "summary": str(source_rows[1][1]).replace("\n", " "),
            "module": str(source_rows[1][2]).replace("\n", " "),
            "validity": "Not a CaseOps item",
            "repo_match": "No Degree Course or Degree Program implementation exists in this repository.",
            "verdict": "Inconclusive",
            "root_cause": "External Edumatica scope. The reported Days/Years mapping cannot be reproduced in CaseOps.",
            "action": "No unrelated code change. Request the Edumatica source and deployed URL for a real fix.",
            "proof": "Repository search: no matching duration-type or degree-program read/write path.",
            "production": "Not tested: external product",
            "notes": "Valid-looking external bug; not a CaseOps bug.",
        },
        {
            "id": "3", "reported_type": "Enhancement",
            "summary": str(source_rows[2][1]).replace("\n", " "),
            "module": str(source_rows[2][2]).replace("\n", " "),
            "validity": "Not a CaseOps item",
            "repo_match": "No matching Degree Course or duration-type dropdown exists in this repository.",
            "verdict": "Inconclusive",
            "root_cause": "External Edumatica scope. Hours cannot be added safely without the owning product code.",
            "action": "No unrelated code change. Request the Edumatica repository or deployment context for implementation.",
            "proof": "Repository search: only unrelated legal time-entry duration fields were found.",
            "production": "Not tested: external product",
            "notes": "Valid-looking external enhancement; not a CaseOps enhancement.",
        },
        {
            "id": "ADJ-REOPEN-2026-07-28", "reported_type": "Bug",
            "summary": "Case reactivation through pre-reopen conflict clearance on deployed production",
            "module": "CaseOps Matter lifecycle / conflict gate",
            "validity": "Valid deployed CaseOps finding",
            "repo_match": "Yes. Current branch has lifecycle-version and reopen-audit guards.",
            "verdict": "Inconclusive",
            "root_cause": "Local candidate rejects stale clearance, but production accepted it with HTTP 200. This is deployed-build drift or an unshipped lifecycle fix.",
            "action": "Keep the lifecycle state machine and registered production regression. Deploy the candidate, expose/prove build identity, and rerun the same spec.",
            "proof": "Local Playwright 3/3 passed; local API stale-clearance test passed; production spec ram-2026-07-15-prod.spec.ts:676 failed with HTTP 200.",
            "production": "Failed on 2026-07-28",
            "notes": "Do not mark fixed until deployed proof passes.",
        },
        {
            "id": "ADJ-NOTICE-FILTER-2026-07-28", "reported_type": "Suspected bug",
            "summary": "Combined notice search/status/owner filter briefly returned an empty UI state on production",
            "module": "CaseOps global Notice Management",
            "validity": "Unconfirmed production flake",
            "repo_match": "Yes. Notice register and filter endpoints exist.",
            "verdict": "Inconclusive",
            "root_cause": "Not reproduced on retry. Direct authenticated production API probe returned the filtered record; no safe root-cause patch is justified from one run.",
            "action": "Retain the existing local and production regression; monitor for recurrence instead of masking it with a longer wait.",
            "proof": "First production replay showed empty result; second full replay passed Notice workflow; direct API combined filter returned total=1.",
            "production": "Passed on retry; initial run flaky",
            "notes": "Not promoted to a confirmed code defect.",
        },
    ]

    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"
    analysis = wb.create_sheet("Analysis")
    evidence = wb.create_sheet("Test Evidence")
    reopen = wb.create_sheet("Reopen Audit")

    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A6"

    style_title(summary, "Ram 28 Jul 2026 Bug / Enhancement Audit", "Scope checked against CaseOps repository and deployed legal tenant. Passwords omitted. Verdicts are fail-closed.", 10)
    for cell, label in zip(["A4", "C4", "E4", "G4", "I4"], ["Total reviewed", "Enhancements", "Bugs / suspected", "Inconclusive", "Valid CaseOps findings"]):
        summary[cell] = label
        summary[cell].font = Font(name="Aptos", size=10, bold=True, color=MUTED)
    summary["A5"] = "=COUNTA('Analysis'!$A$7:$A$11)"
    summary["C5"] = '=COUNTIF(\'Analysis\'!$B$7:$B$11,"Enhancement")'
    summary["E5"] = '=COUNTIF(\'Analysis\'!$B$7:$B$11,"Bug")+COUNTIF(\'Analysis\'!$B$7:$B$11,"Suspected bug")'
    summary["G5"] = '=COUNTIF(\'Analysis\'!$G$7:$G$11,"Inconclusive")'
    summary["I5"] = '=COUNTIF(\'Analysis\'!$E$7:$E$11,"Valid deployed CaseOps finding")'
    for cell in [summary["A5"], summary["C5"], summary["E5"], summary["G5"], summary["I5"]]:
        cell.font = Font(name="Aptos Display", size=18, bold=True, color=NAVY)
        cell.fill = PatternFill("solid", fgColor=PALE_TEAL)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    summary.row_dimensions[5].height = 34
    summary.append([])
    summary.append(["Area", "Finding", "Decision", "Required next action"])
    style_header_row(summary, 7, 1, 4)
    for row in [
        ["Workbook scope", "All three supplied rows describe Edumatica, not CaseOps.", "Out of scope for CaseOps", "Provide the Edumatica repository or deployment context before implementation."],
        ["Case reopening", "Production accepted a conflict check captured before reopen and reactivated the Matter.", "Inconclusive; production failed", "Deploy the current candidate and rerun the dated production lifecycle spec with build identity."],
        ["Notice filter", "First production run showed an empty filtered state; retry and direct API probe passed.", "Inconclusive; monitor", "Retain regression and investigate only on recurrence with network evidence."],
        ["Verification quality", "Local and production surfaces did not prove the same release behavior.", "Process defect confirmed", "Never issue a clean fix verdict without deployed commit/build proof."],
    ]:
        summary.append(row)
    style_body(summary, 8, 11, 1, 4)
    for col, width in {"A": 22, "B": 58, "C": 26, "D": 58}.items():
        summary.column_dimensions[col].width = width
    summary.auto_filter.ref = "A7:D11"

    headers = ["ID", "Reported Type", "Summary", "Module", "Validity", "Repository Match", "Verdict", "Root Cause / Assessment", "Action Taken", "Regression / Proof", "Production Status", "Notes"]
    style_title(analysis, "Item-by-Item Assessment", "Original workbook rows plus adjacent CaseOps findings discovered during the same audit.", len(headers))
    analysis.append([])
    analysis.append([])
    analysis.append([])
    analysis.append(headers)
    style_header_row(analysis, 6, 1, len(headers))
    keys = ["id", "reported_type", "summary", "module", "validity", "repo_match", "verdict", "root_cause", "action", "proof", "production", "notes"]
    for item in rows:
        analysis.append([item[key] for key in keys])
    style_body(analysis, 7, 11, 1, len(headers))
    for index, width in enumerate([18, 18, 52, 34, 28, 46, 18, 54, 54, 54, 26, 36], start=1):
        analysis.column_dimensions[get_column_letter(index)].width = width
    analysis.freeze_panes = "A7"
    analysis.auto_filter.ref = "A6:L11"
    table = Table(displayName="AuditItems", ref="A6:L11")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    analysis.add_table(table)
    analysis.conditional_formatting.add("G7:G11", FormulaRule(formula=['$G7="Inconclusive"'], fill=PatternFill("solid", fgColor=PALE_AMBER)))

    evidence_headers = ["Date", "Surface", "Exact proof artifact", "Result", "What it proves", "Limitations / caveat"]
    style_title(evidence, "Verification Evidence", "A fix claim is paired with the exact command/spec and the limitation of that evidence.", len(evidence_headers))
    evidence.append([])
    evidence.append([])
    evidence.append([])
    evidence.append(evidence_headers)
    style_header_row(evidence, 6, 1, len(evidence_headers))
    for row in [
        ["2026-07-28", "Local Playwright", "tests/e2e/ram-2026-07-15-bugs.spec.ts", "PASS: 3/3", "Default Active creation, global notices, and full dispose/reopen lifecycle pass in a fresh local legal tenant.", "Local proof is not production proof."],
        ["2026-07-28", "Local API", "test_matter_lifecycle.py::test_reopen_invalidates_pre_reopen_conflict_clearance; test_notices.py::test_multi_matter_notice_filters_assignment_and_updates", "PASS: 2/2", "Backend stale-clearance rejection and notice filter semantics pass.", "SQLite API tests do not prove deployed build identity."],
        ["2026-07-28", "Local web", "app/app/notices/page.test.tsx", "PASS: 16/16", "Notice UI filter request behavior and user-visible state pass with mocked client boundaries.", "Mocked UI test is not end-user deployment proof."],
        ["2026-07-28", "Production Playwright", "tests/e2e/ram-2026-07-15-prod.spec.ts", "FAIL: lifecycle returned HTTP 200", "Production login, Active creation, notice workflow, disposal, and reopen executed; stale pre-reopen clearance was accepted.", "Production build identity unavailable; verdict remains Inconclusive."],
        ["2026-07-28", "Production API probe", "Authenticated combined notice query", "PASS: total=1", "The notice search/status/owner filter returned the newly-created record after the failed UI replay.", "Initial UI empty state was not reproduced; monitor rather than patch blindly."],
        ["2026-07-28", "Browser connector", "In-app browser discovery", "BLOCKED: unavailable", "No in-app browser was available on this workstation.", "Standalone Playwright was used instead; no claim of in-app browser proof."],
    ]:
        evidence.append(row)
    style_body(evidence, 7, 12, 1, len(evidence_headers))
    for index, width in enumerate([16, 24, 62, 24, 62, 52], start=1):
        evidence.column_dimensions[get_column_letter(index)].width = width
    evidence.auto_filter.ref = "A6:F12"
    evidence.freeze_panes = "A7"

    reopen_headers = ["Control", "Finding", "Permanent rule"]
    style_title(reopen, "Why Cases Reopen", "Root-cause analysis of the current CaseOps lifecycle and release-verification failure.", len(reopen_headers))
    reopen.append([])
    reopen.append([])
    reopen.append([])
    reopen.append(reopen_headers)
    style_header_row(reopen, 6, 1, 3)
    for row in [
        ["Authoritative transition", "Cases should reopen only through the dedicated Disposed -> Intake lifecycle endpoint.", "Generic metadata PATCH, background writers, and child updates must fail closed."],
        ["Concurrency", "A stale browser edit must not replay an old status after another session disposes the Matter.", "Compare expected status and updated_at under the parent lock; return 409 on mismatch."],
        ["Reopen safety", "Reopen must not resurrect tasks, deadlines, hearings, reminders, or provider calendar rows.", "Neutralize legacy/open children while the parent is locked and keep external sync tombstones durable."],
        ["Conflict clearance", "The pre-reopen check must not activate a reopened Matter.", "Bind every check to lifecycle_version and reject checks at or before the latest reopen audit event."],
        ["Deployment drift", "Local branch passed the lifecycle regression; production returned 200 for the stale-clearance activation.", "Do not mark fixed until the candidate is deployed, build identity is proven, and the same dated production spec passes."],
        ["Product boundary", "The supplied workbook names Edumatica modules, not CaseOps modules.", "Classify external rows as out of scope; request the owning repository instead of implementing a look-alike feature."],
    ]:
        reopen.append(row)
    style_body(reopen, 7, 12, 1, 3)
    reopen.column_dimensions["A"].width = 24
    reopen.column_dimensions["B"].width = 68
    reopen.column_dimensions["C"].width = 70
    reopen.auto_filter.ref = "A6:C12"
    reopen.freeze_panes = "A7"

    for ws in wb.worksheets:
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_view.zoomScale = 90
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    return OUTPUT


if __name__ == "__main__":
    print(build())
