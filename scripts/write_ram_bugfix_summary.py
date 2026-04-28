"""Write the Ram 2026-04-24 bug-fix summary spreadsheet.

Output: ``C:/Users/mishr/Downloads/CaseOps_BugFix_Ram_2026-04-24.xlsx``

Mirrors the format the user has been using for prior bug-fix
summaries (Hari batches I/II/III): one row per bug, columns for
verdict, root cause, files changed, tests added, commit, follow-up
risk.
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

OUT = Path(r"C:/Users/mishr/Downloads/CaseOps_BugFix_Ram_2026-04-24.xlsx")

ROWS = [
    {
        "bug_id": "BUG-011",
        "severity": "Critical (P1)",
        "module": "Security / Global — CSRF Validation",
        "summary": (
            "'Missing CSRF token' across the entire app — every "
            "mutating request fails."
        ),
        "verdict": "Fixed",
        "root_cause": (
            "Set-Cookie was emitted without an explicit Domain= so the "
            "browser scoped both caseops_session and caseops_csrf to "
            "api.caseops.ai. The web app on caseops.ai cannot read "
            "api.caseops.ai cookies via document.cookie, so its CSRF "
            "echo path read nothing and every mutating request landed "
            "without X-CSRF-Token. The CSRF middleware fail-closed "
            "with 403 'Missing CSRF token' on every form submit."
        ),
        "files_changed": (
            "apps/api/src/caseops_api/core/cookies.py "
            "(_cookie_domain helper + Domain plumbed through both "
            "session and portal session/CSRF cookie issuers + "
            "clearers)"
        ),
        "tests_added": (
            "apps/api/tests/test_auth_cookies.py: "
            "test_cookies_set_parent_domain_in_non_local_env + "
            "test_cookies_omit_domain_in_local_env. "
            "tests/e2e/smoke/bug011-csrf-probe.spec.ts: "
            "real-browser probe that bootstraps a tenant, signs in "
            "via the web origin, and dumps cookie/header state on "
            "every API request so this class never silently regresses."
        ),
        "commit": "89bf4b3",
        "deploy": (
            "caseops-api:89bf4b3 + caseops-web:89bf4b3 in asia-south1"
        ),
        "followup_risk": (
            "If a workspace ever uses a non-caseops.ai domain "
            "(white-label / enterprise install), the operator MUST "
            "set CASEOPS_COOKIE_DOMAIN to the parent of both web and "
            "api hostnames. Documented in the cookie module docstring."
        ),
    },
    {
        "bug_id": "BUG-012",
        "severity": "Critical (P1)",
        "module": "UI / Responsive Layout (Dashboard)",
        "summary": (
            "Dashboard horizontally scrolls on mobile (Android Chrome). "
            "Reopened from the Apr 22 batch."
        ),
        "verdict": "Fixed",
        "root_cause": (
            "The prior reopen patched dialogs (rule 6 in "
            "feedback_root_cause_patterns_2026_04_22.md) but missed "
            "the app shell flex container itself. The outer wrapper "
            "had no overflow-x-hidden and the inner flex column had "
            "no min-w-0, so any wide child (long matter title, "
            "table, dialog grid) pushed the body wider than viewport."
        ),
        "files_changed": (
            "apps/web/app/app/layout.tsx — overflow-x-hidden on the "
            "outer flex, min-w-0 on both the inner flex column and "
            "the main element."
        ),
        "tests_added": (
            "tests/e2e/smoke/prod.spec.ts: 360x800 mobile viewport "
            "test that asserts scrollWidth <= clientWidth+1 on /app. "
            "Would have caught the prior reopen at smoke time."
        ),
        "commit": "89bf4b3",
        "deploy": "caseops-web:89bf4b3",
        "followup_risk": (
            "QG-UI-013 (mobile viewport for every page) is still "
            "tracked as P1 in the strict audit — the smoke now "
            "covers /app dashboard but not every authenticated page. "
            "P1-004 work captures this."
        ),
    },
    {
        "bug_id": "BUG-013",
        "severity": "Critical (P1)",
        "module": "Research / Search API",
        "summary": (
            "Research search returns 403 — POST /api/authorities/search."
        ),
        "verdict": "Fixed (downstream of BUG-011)",
        "root_cause": (
            "Same root cause as BUG-011. POST /api/authorities/search "
            "is a mutating method from CSRF's perspective even though "
            "it semantically reads. With no X-CSRF-Token header on "
            "the request (because the cookie was unreadable on "
            "caseops.ai), the CSRF middleware returned 403 'Missing "
            "CSRF token' before the route ever ran. authorities:search "
            "capability is _ALL_AUTHENTICATED so it would have "
            "succeeded for any signed-in user once the header lands."
        ),
        "files_changed": "(no separate fix — BUG-011's commit closes both)",
        "tests_added": (
            "Covered by the BUG-011 cross-subdomain probe + the "
            "existing test_post_with_matching_csrf_header_is_accepted "
            "(now meaningful in prod again after the cookie fix)."
        ),
        "commit": "89bf4b3",
        "deploy": "caseops-api:89bf4b3",
        "followup_risk": (
            "Consider whether /api/authorities/search should be GET. "
            "It would naturally be CSRF-exempt and would not have "
            "exhibited this bug. Tracked as a follow-up but not part "
            "of this fix because the cookie fix resolves it cleanly."
        ),
    },
]


COLUMNS = [
    ("Bug ID", "bug_id", 12),
    ("Severity", "severity", 14),
    ("Module / Sub-module", "module", 36),
    ("Summary", "summary", 60),
    ("Verdict", "verdict", 22),
    ("Root cause", "root_cause", 80),
    ("Files changed", "files_changed", 60),
    ("Tests added", "tests_added", 80),
    ("Commit", "commit", 12),
    ("Deploy", "deploy", 40),
    ("Follow-up risk", "followup_risk", 60),
]


def main() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "BugFix Summary"

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    body_font = Font(size=10)
    wrap = Alignment(wrap_text=True, vertical="top")

    for col_idx, (label, _, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = wrap
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for row_idx, row in enumerate(ROWS, start=2):
        for col_idx, (_, key, _) in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row[key])
            cell.font = body_font
            cell.alignment = wrap

    ws.row_dimensions[1].height = 24
    for r in range(2, 2 + len(ROWS)):
        ws.row_dimensions[r].height = 180

    ws.freeze_panes = "A2"
    wb.save(OUT)
    print(f"wrote {OUT}")


if __name__ == "__main__":
    main()
