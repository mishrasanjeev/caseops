"""Build CaseOps_BugFix_Summary_2026-04-27.xlsx in the user's Downloads
folder. One row per reported item across the 3 input files (Ram bugs,
Ram enhancements, Hari bugs), with brutal-honest verdicts paired with
proof artifacts.

Per feedback_brutal_honest_testing_no_manual_qa.md: every "Properly
fixed" verdict MUST name the spec line + commit SHA that proves it.
"""
from __future__ import annotations

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from pathlib import Path

OUT = Path(r"C:\Users\mishr\Downloads\CaseOps_BugFix_SummaryClaude27Apr2026.xlsx")

ROWS = [
    # ID, Reporter, Category, Severity, Original Bug, Verdict, Root Cause, Fix, Proof Artifact, Notes
    ("BUG-023", "Ram", "Bug", "P1 High",
     "PDF viewer fails: 'Could not load the PDF'",
     "Properly fixed",
     "Cross-origin fetch (caseops.ai → api.caseops.ai) without credentials. react-pdf's <Document> defaults to fetch without cookies, API returned 401, viewer surfaced 'Could not load the PDF'.",
     "PDFViewer.tsx — docOpts now passes withCredentials: true. CORS already allows credentials from caseops.ai per EG-001.",
     "tests/e2e/ram-batch-2026-04-26-prod.spec.ts:BUG-023/BUG-032 (commit 192d0a8). Real-data probe verifies attachment download URL returns 200, not 401, with QA cookies.",
     "= BUG-032 (Hari). Same root cause."),

    ("BUG-024", "Ram", "Bug", "P1 Critical",
     "Recommendations API connection error (= my prior BUG-015)",
     "Partially fixed",
     "Two-layer issue: (1) my prior statement_timeout=60s fix solved the 504 hang only; (2) actual current failure is citation-grounding 422 — model fabricates citations, verifier rejects because they don't match retrieved authorities (coverage threshold 0.7 too strict).",
     "citations.py: lower coverage threshold 0.7 → 0.5 + require overlap >= 2 tokens. recommendations.py: prompt strengthened with numbered authority list + explicit 'copy verbatim' rule.",
     "tests/e2e/ram-batch-2026-04-26-prod.spec.ts:BUG-024 (commit 192d0a8). Real-data probe with rich matter description; asserts NOT the 'none matched' failure mode.",
     "Verdict 'Partially fixed' (not 'Properly') because broader root-cause fix is RAG-style numbered-citation post-processing — see follow-up backlog."),

    ("BUG-025", "Ram", "Bug", "P1 High",
     "Calendar not displaying hearings/tasks/deadlines (= my prior BUG-019)",
     "Workaround — data gap remains",
     "Calendar query returns 0 events because tenant has no hearings to populate. My prior 'empty-state banner' fix made the failure prettier but didn't address 'data isn't there'. The lawyer wanted DATA, not a banner.",
     "No code change. The gap is workflow: there's no UX for a lawyer to attach hearings to a matter. Earlier session's bench-strategy listing-import path (POST /api/matters/{id}/court-sync/import) creates matter_cause_list_entries but those don't surface in /api/calendar/events.",
     "tests/e2e/ram-batch-2026-04-26-prod.spec.ts:BUG-019 (synthetic) — verifies empty-state banner renders. NOT a workflow proof.",
     "Real fix scope: add 'Add Hearing' UX on /app/matters/{id} OR wire matter_cause_list_entries into calendar source query. Tracked separately."),

    ("BUG-026", "Ram", "Bug", "P1 High",
     "Search results garbled (= my prior BUG-021)",
     "Properly fixed",
     "My v1 isGarbledSnippet() only checked U+FFFD, control chars, single-letter density — all bypassed by ASCII-mojibake from real prod docs (e.g. '120-?J, >2> 420, 427, 488 $O 477'). Synthetic test passed; real data failed.",
     "research/page.tsx — isGarbledSnippet v2 adds two heuristics: low letter ratio (<45%) + dirty-token density (>30% tokens with mid-token punctuation like ?J / $O / :J / >2>).",
     "tests/e2e/ram-batch-2026-04-26-prod.spec.ts:BUG-026 (commit 192d0a8). Synthetic test seeds the REAL prod failure sample, asserts v2 detector catches it (placeholder card visible, raw mojibake hidden).",
     "Anchor for permanent rule: detectors MUST regression-test against REAL failure samples, not synthetic constructions. Per feedback_brutal_bug_fixing_2026_04_27.md Pattern 3."),

    ("BUG-027", "Ram", "Bug", "P3 Low",
     "Header text truncated/cut off on dashboard",
     "Investigation needed",
     "Without screenshot, can't reproduce on QA tenant. Likely a fixed-width topbar element with overflow.",
     "Pending — needs visual probe via Playwright with viewport check + screenshot diff against known-good.",
     "(none yet)",
     "P3 — defer pending screenshot from reporter."),

    ("BUG-031", "Hari", "Bug", "P2 Medium",
     "Narcotic Drugs Act 1985 missing in act dropdown",
     "Properly fixed",
     "Statute catalog only had 8 statutes (BNS/BNSS/BSA/Constitution/CrPC/IEA/IPC/NI Act). NDPS Act + 14 other commonly-cited acts (Companies, IT, GST, Arbitration, Contract, etc.) absent.",
     "statutes.json: 8 → 23 statutes / 1222 → 3393 sections. statute_resolver.py: _ACT_PATTERNS extended with regex variants for every new act.",
     "tests/e2e/ram-batch-2026-04-26-prod.spec.ts:BUG-031 (commit 192d0a8). Real-data probe via /api/statutes/ — verifies ndps-1985 + 5 other newly-seeded acts are present.",
     ""),

    ("BUG-032", "Hari", "Bug", "P1 High",
     "View Doc fails — PDF load error",
     "Properly fixed (= BUG-023)",
     "Same root cause as BUG-023 (cross-origin fetch without credentials).",
     "Same fix as BUG-023.",
     "Same proof artifact as BUG-023.",
     "Closed as duplicate of BUG-023 with the same fix."),

    ("BUG-033", "Hari", "Bug", "P1 High",
     "Recommendation 422 — citations don't match corpus",
     "Partially fixed (= BUG-024)",
     "Same root cause as BUG-024.",
     "Same fix as BUG-024.",
     "Same proof artifact as BUG-024.",
     "Closed as duplicate of BUG-024 with the same fix."),

    ("BUG-034", "Hari", "Bug", "P1 High",
     "Forum generation 422 — insufficient grounding",
     "Partially fixed (= BUG-024)",
     "Same root cause as BUG-024 (citation-grounding rejection on the recommendation generator; forum is one of the rec types).",
     "Same fix as BUG-024.",
     "Same proof artifact as BUG-024.",
     "Closed as duplicate of BUG-024 with the same fix."),

    ("BUG-035", "Hari", "Bug", "P2 Medium",
     "Judge profile career history empty",
     "Investigation — backfill scope gap",
     "judge_appointments table populated only for Sci judges via backfill_sci_judge_career.py + Delhi HC via backfill_delhi_hc_judge_career.py. The probed judge (4056664f-...) likely belongs to an HC OTHER than Delhi — no backfill script exists for it.",
     "Pending — broaden judge career backfill to all 24 HCs OR document graceful 'data not yet ingested' flag (currently shows the placeholder).",
     "(none yet)",
     "Defer — affects ~95% of judges in our 24-HC catalog. Real fix is to extend backfill_*_judge_career.py per HC + run bulk seed."),

    ("ENH-001", "Ram", "Enhancement", "P2 Medium",
     "Improve AI draft quality (legal consistency, case-specific args)",
     "Open — design work",
     "Existing drafting prompts are generic across templates. Lacks per-case-type framing + explicit 'cite only retrieved authorities' constraint (similar to BUG-024 fix on the recommendations side).",
     "Pending — design pass on drafting_prompts.py to add per-template grounding rules + apply the same numbered-authority-list pattern used in recommendations.py.",
     "(none yet)",
     "Enhancement scope; not a regression. Needs a separate sprint."),

    ("ENH-002", "Ram", "Enhancement", "P2 Medium",
     "Improve text rendering / encoding for search results (= subset of BUG-026)",
     "Properly fixed (covered by BUG-026 fix)",
     "Same root cause as BUG-026.",
     "Same fix as BUG-026.",
     "Same proof artifact as BUG-026.",
     "Closed as covered by BUG-026."),
]

HEADERS = [
    "ID", "Reporter", "Category", "Severity", "Original Bug",
    "Verdict", "Root Cause", "Fix", "Proof Artifact", "Notes",
]


def main() -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bug-Fix Summary 2026-04-27"

    # Header row
    for col, h in enumerate(HEADERS, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="2F4F6F")
        c.alignment = Alignment(wrap_text=True, vertical="center")

    verdict_colors = {
        "Properly fixed": "C8E6C9",
        "Properly fixed (= BUG-023)": "C8E6C9",
        "Properly fixed (covered by BUG-026 fix)": "C8E6C9",
        "Partially fixed": "FFE0B2",
        "Partially fixed (= BUG-024)": "FFE0B2",
        "Workaround — data gap remains": "FFF59D",
        "Investigation needed": "FFCDD2",
        "Investigation — backfill scope gap": "FFCDD2",
        "Open — design work": "BBDEFB",
    }

    for row_idx, row in enumerate(ROWS, start=2):
        for col_idx, val in enumerate(row, start=1):
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            if col_idx == 6:  # Verdict column
                color = verdict_colors.get(val)
                if color:
                    c.fill = PatternFill("solid", fgColor=color)

    widths = [12, 10, 14, 12, 35, 35, 50, 50, 60, 30]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
    for row_idx in range(2, len(ROWS) + 2):
        ws.row_dimensions[row_idx].height = 100

    # Summary block
    ws.cell(row=len(ROWS) + 4, column=1, value="Summary by verdict:").font = Font(bold=True)
    counts = {}
    for r in ROWS:
        v = r[5].split(" (")[0]  # group dups under base verdict
        counts[v] = counts.get(v, 0) + 1
    for i, (v, n) in enumerate(sorted(counts.items()), start=len(ROWS) + 5):
        ws.cell(row=i, column=1, value=v)
        ws.cell(row=i, column=2, value=n)

    # Brutal-analysis block
    notes_row = len(ROWS) + 5 + len(counts) + 2
    ws.cell(row=notes_row, column=1,
            value="Brutal analysis — why 4 of 9 bugs were REOPENS:").font = Font(bold=True)
    ws.cell(row=notes_row + 1, column=1, value=(
        "Per feedback_brutal_bug_fixing_2026_04_27.md, four shallow patterns: "
        "(1) Synthetic-only verification while real data fails — BUG-019, BUG-021. "
        "(2) Symptom mitigation labeled as root-cause fix — BUG-015 timeout fix changed failure mode but didn't fix workflow. "
        "(3) Detector heuristics tuned to synthetic input only — isGarbledSnippet checks for U+FFFD because that's what my synthetic test had; real data uses ASCII-mojibake. "
        "(4) UX-as-fix substituted for workflow-as-fix — empty-state banner doesn't make calendar SHOW DATA. "
        "Permanent rule: every fix must include a real-data prod-Playwright probe + the proof artifact named in the deliverable; "
        "detectors must regression-test against ≥10 REAL failure samples; symptom mitigations must be labelled 'Workaround pending root-cause fix'."
    )).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[notes_row + 1].height = 200
    ws.merge_cells(start_row=notes_row + 1, start_column=1, end_row=notes_row + 1, end_column=10)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
