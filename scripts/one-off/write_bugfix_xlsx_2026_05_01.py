"""One-off: write the Ram 2026-05-01 bug-fix summary xlsx.

Lives under scripts/one-off because it's not a recurring tool — just
the same script we run after every Ram batch. Output path:
``C:/Users/mishr/Downloads/CaseOps_BugFix_Summary_Ram01May2026.xlsx``.
"""
from __future__ import annotations

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


def main() -> int:
    wb = openpyxl.Workbook()

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell_align = Alignment(horizontal="left", vertical="top", wrap_text=True)
    thin = Side(border_style="thin", color="CCCCCC")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    verdict_fills = {
        "Properly fixed": PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid"),
        "Partially fixed": PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
        "Not fixed": PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
        "Inconclusive": PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid"),
    }

    # Sheet 1
    ws = wb.active
    ws.title = "Bug-fix Summary 2026-05-01"

    headers = [
        "ID", "Type", "Severity", "Module", "Summary (as reported)",
        "Status (Ram)", "Triage", "Root cause", "Fix shipped (commit / file)",
        "Verdict", "Verification", "Regression test", "Reopen pattern",
    ]
    ws.append(headers)
    for col, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    rows = [
        [
            "BUG-028", "Functional", "High",
            "Documents - PDF Viewer",
            "PDF document is not loading in viewer; shows 'Could not load the PDF. Try the direct download above.'",
            "ReOpen (3rd reopen)",
            (
                "VALID BUG. Same surface error as BUG-023 / BUG-032 (which fixed the cross-origin "
                "cookie path). Different root cause this time: the PDF.js worker URL pointed at "
                "https://unpkg.com/... and prod CSP `worker-src 'self' blob:` blocked it. Worker "
                "fails to spawn -> react-pdf surfaces the generic error UI."
            ),
            (
                "PDF.js workerSrc bundled via the `new URL('pdfjs-dist/build/pdf.worker.min.mjs', "
                "import.meta.url)` Webpack/Turbopack idiom - Next.js emits the worker at "
                "/_next/static/media/pdf.worker.<hash>.mjs (same-origin, CSP passes). pdfjs-dist "
                "is a transitive dep of react-pdf so no new package needed."
            ),
            "Commit 06f63a9 - apps/web/components/document/PDFViewer.tsx",
            "Properly fixed",
            (
                "Prod-Playwright tests/e2e/ram-batch-2026-05-01-prod.spec.ts -> 'PDF viewer does NOT "
                "show Could not load the PDF (CSP-safe worker)'. Asserts (a) the error UI text has "
                "count 0 and (b) a <canvas> element renders (proving the worker spawned + page rendered)."
            ),
            (
                "tests/e2e/ram-batch-2026-05-01-prod.spec.ts (added). Wired into "
                "playwright.prod-ram.config.ts testMatch."
            ),
            (
                "Pattern 1 (Brutal 2026-05-01): generic surface error -> assumed same root cause as "
                "prior reopen. Same UI string covered TWO different bugs; I declared victory after "
                "fixing only one. Hard rule added: never trust local repro on a happy-path test "
                "asset for a reopened bug - pull prod console first."
            ),
        ],
        [
            "BUG-029", "Functional / API", "Critical",
            "Recommendations - Authority generation",
            (
                "Recommendation generation fails with 502 due to invalid JSON response from LLM "
                "(GPT-5.1). 'LLMResponseFormatError: model did not return valid JSON'."
            ),
            "ReOpen (vs BUG-035 on 2026-04-28)",
            (
                "VALID BUG. BUG-035 widened the _LLMResponse schema bounds (rationale 6000->15000, "
                "options max 5->10, etc). That fix worked when GPT was returning valid JSON that "
                "just exceeded old size limits. The reopen is a different failure mode: GPT-5.1 "
                "occasionally emits actual malformed JSON (~1-2% of long structured outputs). With "
                "the Anthropic->Haiku->OpenAI fallback ladder removed in this session, a single "
                "transient flake = 502."
            ),
            (
                "Single retry on LLMResponseFormatError specifically - same provider, same model, "
                "same prompt. Output is non-deterministic at temperature > 0 so most format errors "
                "clear on retry. Quota / 5xx / timeout still 502 immediately (retry won't fix "
                "upstream outages). Pattern audit applied the same retry shape to "
                "services/drafting.py + services/hearing_packs.py (both had the same anti-pattern)."
            ),
            (
                "Commit 06f63a9 - apps/api/src/caseops_api/services/recommendations.py + "
                "services/drafting.py + services/hearing_packs.py"
            ),
            "Properly fixed",
            (
                "Backend pytest tests/test_recommendations.py::"
                "test_recommendation_format_error_retries_once_then_succeeds (asserts call_count == 2 "
                "+ status 200). And test_recommendation_format_error_retry_also_fails_yields_502 "
                "(asserts the retry is bounded - exactly 1 retry then 502). Plus prod-Playwright "
                "spec asserting recs endpoint is reachable and not 502 on the QA Bot session."
            ),
            (
                "apps/api/tests/test_recommendations.py (2 new cases) + "
                "tests/e2e/ram-batch-2026-05-01-prod.spec.ts."
            ),
            (
                "Pattern 2 (Brutal 2026-05-01): removed the multi-provider fallback ladder in the "
                "same session without adding the explicit same-provider retry. Lost the safety net "
                "that was masking transient GPT format errors. Hard rule added: removing a fallback "
                "ladder requires shipping the explicit retry path in the same commit."
            ),
        ],
        [
            "ENH-003", "Enhancement", "Medium",
            "Drafting - AI Draft Generator",
            (
                "Improve AI-generated draft quality with stronger legal reasoning, case-specific "
                "arguments, structured formatting (Facts / Grounds / Prayer)."
            ),
            "Open",
            (
                "PARTIALLY ADDRESSED PRE-REPORT. Sprints 1-7 of the 12-sprint drafting roadmap "
                "(shipped 2026-05-01 prior to this batch) added: (a) 20 specialised templates with "
                "per-template prompts enforcing required structure, (b) bench-aware drafting "
                "injecting the assigned judge's recurring tests + practice-area patterns + cited "
                "authorities into the prompt, (c) court-format-aware PDF export so structured drafts "
                "come out with the right margins / page numbers / court header. Ram likely tested the "
                "legacy generic 'Brief / Notice / Reply' path (see ENH-004 fix) where none of these "
                "apply - hence the 'generic arguments, lacks case-specific depth' observation."
            ),
            (
                "No new code change for ENH-003 itself - Sprints 1-7 already address this. ENH-004 "
                "fix (this batch) makes those sprints DISCOVERABLE so Ram can actually try a "
                "template-based draft."
            ),
            (
                "Sprints 1-7 commits: a1f8612 (Sprint 1 - 4 templates), 164c4f6 (Sprint 2 - 7 more), "
                "d6207e0 (Sprint 3 - court PDF), df71b4e (Sprint 4 - filing bundle), 22f9883 "
                "(Sprint 5 - 6 more court profiles), 9a316bd (Sprint 6 - revision diff), 1a2662a "
                "(Sprint 7 - bench-aware expanded to 15 templates). Plus 06f63a9 (this batch - "
                "make those discoverable via ENH-004 fix)."
            ),
            "Partially fixed",
            (
                "Recommended re-test: with ENH-004 fix live, Ram should generate a writ_petition / "
                "quashing_petition / civil_suit draft on a matter with a known judge. Expected: "
                "structured Facts / Grounds / Prayer body, citations bracketed, bench-history "
                "paragraph if the judge has an indexed history."
            ),
            (
                "tests/test_drafting_bench_aware.py - 42 parameterised cases assert the bench-aware "
                "injection fires for 15 of 20 templates; tests/test_drafting_templates.py - 11 facts-"
                "validation + 11 prompt-correctness cases. Live-LLM eval harness is Sprint 12."
            ),
            (
                "Not a reopen. The reason this looked like a regression to Ram is Pattern 3 (Brutal "
                "2026-05-01): the backend was complete + the frontend entry-point (New draft button) "
                "didn't lead there. Discoverability is part of the feature."
            ),
        ],
        [
            "ENH-004", "Enhancement (actually a bug)", "Medium",
            "Drafting - Drafting Citation",
            (
                "Add commonly used legal draft types (Bail, Anticipatory Bail, Petition, Affidavit, "
                "Legal Notice, Complaint) - currently only generic options (Brief, Notice, Reply, "
                "etc.) available."
            ),
            "Open",
            (
                "VALID BUG mis-labelled as enhancement. Sprints 1+2 already shipped 20 specialised "
                "templates: bail / anticipatory_bail / writ_petition / quashing_petition / "
                "dv_quashing_petition / civil_suit / written_statement / reply_counter_affidavit / "
                "appeal_memorandum / arbitration_section_9 / criminal_complaint / "
                "amendment_of_pleadings / divorce_petition / compromise_petition / probate_petition / "
                "cheque_bounce_notice / property_dispute_notice / affidavit / vakalatnama / "
                "caveat_petition. They're 100% reachable from /app/matters/{id}/drafts/new. But on "
                "/app/matters/{id}/drafts (the natural starting page), the 'New draft' button "
                "opened a dialog with only 5 legacy types - bypassing the template grid entirely. "
                "Compounding bug: even on /drafts/new, the KNOWN_TEMPLATE_TYPES allow-list was a "
                "9-entry hardcoded set; clicking any of the 11 templates added in Sprints 1+2 was "
                "a silent no-op."
            ),
            (
                "(1) Replaced /drafts/page.tsx 'New draft' dialog with a Link routing to /drafts/new "
                "(the template grid). The grid is now the single source of truth for available "
                "templates. (2) Replaced /drafts/new/page.tsx KNOWN_TEMPLATE_TYPES hardcoded set "
                "with a regex-based gate (snake_case validity check). DraftingStepper itself fetches "
                "the schema and 404s gracefully if the type is bogus, so a permissive gate just "
                "lets the stepper show its own error."
            ),
            (
                "Commit 06f63a9 - apps/web/app/app/matters/[id]/drafts/page.tsx + "
                "apps/web/app/app/matters/[id]/drafts/new/page.tsx"
            ),
            "Properly fixed",
            (
                "Prod-Playwright tests/e2e/ram-batch-2026-05-01-prod.spec.ts -> 'New draft button "
                "routes to template grid'. Asserts (a) URL changes to /drafts/new on click (not a "
                "dialog), (b) at least 10 template cards visible in the grid, (c) clicking "
                "writ_petition card lands on the stepper with type=writ_petition (proving the gate "
                "let the click through)."
            ),
            "tests/e2e/ram-batch-2026-05-01-prod.spec.ts.",
            (
                "Pattern 3 (Brutal 2026-05-01): backend feature complete + frontend entry-point "
                "never clicked in a real browser. Sprints 1-7 were two weeks of work; from Ram's "
                "perspective the only way in was a deep link he didn't have. Hard rule added: "
                "before declaring a feature done, sign in as the target role + click the natural "
                "entry-point button. Plus a derived rule on hardcoded frontend allow-lists "
                "mirroring backend enums."
            ),
        ],
    ]

    for r in rows:
        ws.append(r)
        row_num = ws.max_row
        verdict = r[9]
        fill = verdict_fills.get(verdict)
        for col_idx, _ in enumerate(r, start=1):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.alignment = cell_align
            cell.border = border
            if col_idx == 10 and fill:
                cell.fill = fill
                cell.font = Font(bold=True)

    widths = {
        "A": 12, "B": 22, "C": 10, "D": 28, "E": 50, "F": 22,
        "G": 70, "H": 70, "I": 50, "J": 18, "K": 60, "L": 50, "M": 70,
    }
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w
    ws.row_dimensions[1].height = 30
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 240

    # Sheet 2 - Brutal Analysis
    ws2 = wb.create_sheet("Brutal Analysis 2026-05-01")
    ws2["A1"] = "Three more shallow patterns Ram caught (BUG-028 + BUG-029 + ENH-004)"
    ws2["A1"].font = Font(bold=True, size=14)
    ws2.merge_cells("A1:C1")

    ws2["A3"] = "Pattern"
    ws2["B3"] = "Anchor incident"
    ws2["C3"] = "Hard rule going forward"
    for c in ("A3", "B3", "C3"):
        ws2[c].font = header_font
        ws2[c].fill = header_fill
        ws2[c].alignment = header_align
        ws2[c].border = border

    patterns = [
        [
            "1. Same surface error + assume-same-root-cause",
            (
                "BUG-028 PDF viewer 'Could not load the PDF' was the same UI string as BUG-023 + "
                "BUG-032. I fixed the cross-origin cookie issue (which was real) and assumed I'd "
                "fixed both bugs. The actual root cause this time was CSP blocking the unpkg.com "
                "worker URL - completely unrelated to cookies."
            ),
            (
                "For any reopened bug whose surface message is generic ('Could not load X', "
                "'Something went wrong', 'Failed to generate'), pull the actual prod console / "
                "network error BEFORE designing a fix. Don't trust local repro on a happy-path "
                "test asset. The error message is a SYMPTOM, not the diagnosis. A 'Properly fixed' "
                "verdict for a reopened bug requires a Playwright probe against prod with the QA "
                "Bot session that asserts the EXACT error UI is no longer reachable."
            ),
        ],
        [
            "2. Removing a fallback ladder without adding the retry path",
            (
                "This session I removed the Anthropic->Haiku->OpenAI fallback ladder ('stick to "
                "OpenAI, no fallbacks'). Token spend dropped 3x - but the ladder's accidental "
                "function was masking transient GPT JSON-format errors. Without it, a single 1-2% "
                "format flake = 502 for the user. I shipped the no-fallback change with one "
                "regression test that LITERALLY proved the user-facing break ('primary error -> "
                "502') was now reachable, then declared the change shipped."
            ),
            (
                "When removing a multi-layer error-handling path, ask: 'what was the layer "
                "compensating for, even by accident?' If it was masking transient errors, replace "
                "the layer with explicit retry on the specific transient class - same-provider "
                "retry on LLMResponseFormatError. Categorical fallback removed; explicit retry "
                "stays. Ship both in the same commit."
            ),
        ],
        [
            "3. Backend complete + frontend entry-point never clicked",
            (
                "Sprints 1+2 added 20 specialised drafting templates with full backend tests + "
                "per-template prompts + recommender integration. From Ram's perspective the work "
                "was invisible because the natural entry-point ('New draft' button on "
                "/app/matters/{id}/drafts/) opened a 5-option legacy dialog instead of the template "
                "grid. Compounding: even on the template grid, KNOWN_TEMPLATE_TYPES was a stale "
                "9-entry hardcoded set so 11 templates were silent no-ops on click."
            ),
            (
                "Before declaring a feature done, sign in as the target user role, navigate from "
                "the natural starting page, and click the natural entry-point button. If the "
                "feature requires a deep link to reach, it's not done. Plus: any frontend allow-"
                "list / enum that mirrors a backend list is a future bug - derive from the API "
                "response or remove the gate and let backend errors surface organically."
            ),
        ],
        [
            "Bonus rule on CSP + CDN-loaded assets",
            (
                "Default react-pdf installation pattern uses unpkg.com CDN for the PDF.js worker. "
                "CaseOps prod CSP is locked down (worker-src 'self' blob:; script-src 'self' + "
                "analytics-only). Any third-party CDN asset is automatically blocked."
            ),
            (
                "In a CSP-strict app, never load runtime assets from third-party CDNs (unpkg, "
                "jsdelivr, cdnjs). Use the bundler's URL-import idiom (`new URL('package/path', "
                "import.meta.url)`) so the asset ships from the same origin as the rest of the "
                "build output. Audit every <script src=...> and every dynamic asset URL against "
                "the deployed CSP."
            ),
        ],
    ]

    for p in patterns:
        ws2.append(p)
        rn = ws2.max_row
        for col_idx in range(1, 4):
            cell = ws2.cell(row=rn, column=col_idx)
            cell.alignment = cell_align
            cell.border = border

    ws2.column_dimensions["A"].width = 50
    ws2.column_dimensions["B"].width = 80
    ws2.column_dimensions["C"].width = 90
    ws2.row_dimensions[1].height = 30
    ws2.row_dimensions[3].height = 25
    for rn in range(4, ws2.max_row + 1):
        ws2.row_dimensions[rn].height = 220

    # Sheet 3 - Pattern Audit
    ws3 = wb.create_sheet("Pattern Audit")
    ws3["A1"] = "Repo-wide audit for the same patterns (2026-05-01)"
    ws3["A1"].font = Font(bold=True, size=14)
    ws3.merge_cells("A1:D1")

    ws3.append([])
    ws3.append(["Pattern", "Location checked", "Status", "Notes"])
    for c_idx in range(1, 5):
        cell = ws3.cell(row=3, column=c_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    audit = [
        ["Format-error retry on generate_structured", "services/recommendations.py", "FIXED in 06f63a9", "Single retry on LLMResponseFormatError"],
        ["Format-error retry on generate_structured", "services/drafting.py", "FIXED in 06f63a9", "Parallel pattern fix"],
        ["Format-error retry on generate_structured", "services/hearing_packs.py", "FIXED in 06f63a9", "Parallel pattern fix"],
        ["Format-error retry on generate_structured", "services/contract_intelligence.py", "OK pre-existing", "Already retries on LLMProviderError parent"],
        ["Format-error retry on generate_structured", "services/corpus_structured.py", "OK pre-existing", "Already handles LLMResponseFormatError"],
        ["Format-error retry on generate_structured", "services/corpus_title_reextract.py", "OK pre-existing", "Already handles LLMResponseFormatError"],
        ["Format-error retry on generate_structured", "services/drafting_preview.py", "N/A", "Uses raw provider.generate(), not generate_structured - no JSON validation step to retry on"],
        ["CDN-loaded asset under strict CSP", "components/document/PDFViewer.tsx (pdfjs worker)", "FIXED in 06f63a9", "Bundled via new URL(..., import.meta.url)"],
        ["CDN-loaded asset under strict CSP", "next.config.ts script-src list", "OK", "CSP allows googletagmanager only - no other third-party scripts in the bundle"],
        ["CDN-loaded asset under strict CSP", "Other components for fonts / icons / charts", "OK", "Lucide icons + project fonts ship from /_next/static; no third-party CDN"],
        ["Frontend allow-list mirroring backend enum", "/app/matters/{id}/drafts/new/page.tsx KNOWN_TEMPLATE_TYPES", "FIXED in 06f63a9", "Replaced with permissive regex"],
        ["Frontend allow-list mirroring backend enum", "schemas.ts DraftType enum", "OK (intentional)", "5-value legacy draft_type, distinct from template_type - these are separate fields on the Draft row"],
        ["Backend feature -> not clicked in browser", "Sprints 1+2 templates discoverability", "FIXED in 06f63a9", "ENH-004 root-cause fix"],
        ["Backend feature -> not clicked in browser", "Sprints 3-8 (PDF / bundle / compare / checklist) discoverability", "OK", "All accessed from the draft detail page; shipped with web wiring in the same commit"],
    ]
    for row in audit:
        ws3.append(row)
        rn = ws3.max_row
        for col_idx in range(1, 5):
            cell = ws3.cell(row=rn, column=col_idx)
            cell.alignment = cell_align
            cell.border = border

    ws3.column_dimensions["A"].width = 40
    ws3.column_dimensions["B"].width = 50
    ws3.column_dimensions["C"].width = 22
    ws3.column_dimensions["D"].width = 70

    out = r"C:\Users\mishr\Downloads\CaseOps_BugFix_Summary_Ram01May2026.xlsx"
    wb.save(out)
    print(f"Wrote {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
