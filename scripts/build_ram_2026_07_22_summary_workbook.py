"""Build the Ram 2026-07-22 CaseOps bug-fix evidence workbook.

The source workbook contains tester credentials.  This generated artifact
records which tenant/account was exercised, but deliberately never stores a
password.  Run-time evidence must be updated below before the workbook is
regenerated for delivery.
"""
from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


DEFAULT_OUTPUT = Path(r"C:\tmp\CaseOps_BugFix_Summary_Ram22Jul2026.xlsx")
SOURCE_WORKBOOK = Path(r"C:\Users\mishr\Downloads\CaseOps_Bugs_Ram22Jul2026.xlsx")

BASELINE_COMMIT = "33bf177b45eeb556ea5a02a82f2d8644cd7e5c25"
DEPLOYED_COMMIT = "34f19ad2bc0a5b48398144998cf546cc9e7a815a"
API_REVISION = "caseops-api-00210-fnv"
API_DIGEST = "sha256:23d2e9313cf8a99f538e3dbd5f9a9cfc0533e0559de0fc16f4b02df4a18e3b94"
WEB_REVISION = "caseops-web-00189-k9f"
WEB_DIGEST = "sha256:7ffd1277b78d352539e0a4eeef83e320b3a396227b0c7ad3128f123ba4f15745"
MIGRATION_EXECUTION = "caseops-migrate-job-ggqwz"
INDEPENDENT_QA_RUN = "29929098217"

NAVY = "17324D"
TEAL = "087E8B"
PALE_TEAL = "DDF4F2"
PALE_BLUE = "EAF1F8"
PALE_AMBER = "FFF2CC"
PALE_GREEN = "E2F0D9"
PALE_RED = "FCE8E6"
WHITE = "FFFFFF"
INK = "1F2937"
MUTED = "52606D"
LINE = "C9D4DF"

THIN = Side(style="thin", color=LINE)
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")


def _title(ws, title: str, subtitle: str, last_column: int) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_column)
    ws["A1"] = title
    ws["A1"].font = Font(size=18, bold=True, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_column)
    ws["A2"] = subtitle
    ws["A2"].font = Font(size=10, italic=True, color=MUTED)
    ws["A2"].alignment = WRAP
    ws.row_dimensions[2].height = 30


def _header_row(ws, row: int, columns: int) -> None:
    for cell in ws[row][:columns]:
        cell.fill = PatternFill("solid", fgColor=TEAL)
        cell.font = Font(bold=True, color=WHITE)
        cell.alignment = WRAP
        cell.border = CELL_BORDER
    ws.row_dimensions[row].height = 30


def _table(ws, name: str, start_row: int, end_row: int, end_column: int) -> None:
    ref = f"A{start_row}:{get_column_letter(end_column)}{end_row}"
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def _finish_sheet(ws, *, widths: list[int], landscape: bool = True) -> None:
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width
    for row in ws.iter_rows(min_row=3):
        for cell in row:
            cell.alignment = WRAP
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"
    ws.page_setup.orientation = "landscape" if landscape else "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.oddFooter.center.text = "CaseOps | Ram 22 Jul 2026 | Confidential QA evidence"
    ws.oddFooter.right.text = "Page &P of &N"


def build_executive_summary(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "Executive Summary"
    _title(
        ws,
        "CaseOps — Ram 22 July 2026 issue assessment and fix summary",
        "Formal verdict rule satisfied: the exact candidate commit was deployed and the committed production Playwright regression passed; BUG-001 is now Properly fixed.",
        2,
    )

    metadata = [
        ("Source workbook", str(SOURCE_WORKBOOK)),
        ("Assessment date", "2026-07-22"),
        ("Baseline → deployed runtime commit", f"{BASELINE_COMMIT} → {DEPLOYED_COMMIT}"),
        ("Reported issues", "=COUNTA('Issue Assessment'!A5:A1000)"),
        ("Valid product-policy enhancements", '=COUNTIF(\'Issue Assessment\'!F5:F1000,"Valid enhancement")'),
        ("Valid regressions", '=COUNTIF(\'Issue Assessment\'!F5:F1000,"Valid regression")'),
        (
            "Formal verdict",
            f"Properly fixed — exact candidate commit {DEPLOYED_COMMIT} is deployed; the committed production Playwright specification passed 2/2 with terminal cleanup.",
        ),
        ("Tester identity", "legal / hari.gupta@gmail.com (password supplied at runtime; not stored)"),
    ]
    ws.append([])
    for label, value in metadata:
        ws.append([label, value])
        ws.cell(ws.max_row, 1).font = Font(bold=True, color=NAVY)
        ws.cell(ws.max_row, 1).fill = PatternFill("solid", fgColor=PALE_BLUE)
        ws.cell(ws.max_row, 1).border = CELL_BORDER
        ws.cell(ws.max_row, 2).border = CELL_BORDER
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 115

    start = ws.max_row + 2
    ws.cell(start, 1, "Decision")
    ws.cell(start, 2, "Evidence-backed conclusion")
    _header_row(ws, start, 2)
    decisions = [
        (
            "Classification",
            "BUG-001 is a valid product-policy enhancement, not a regression against the pre-22-Jul contract. The July 15 acceptance explicitly made direct creation Active by default while retaining the Intake/On Hold → Active conflict gate.",
        ),
        (
            "Why it appeared to reopen",
            "The earlier acceptance boundary covered creation, not every route into Active. The old split policy was duplicated in service logic, a gate evaluator, UI recovery copy, lifecycle/reopen tests, PRD/current guidance, and production regressions. Those tests protected the old requirement, so they could not detect the new policy expectation.",
        ),
        (
            "Depth of correction",
            "The candidate removes the server gate and dead evaluator, removes UI blocking guidance, keeps conflict review usable as an advisory workflow, preserves terminal lifecycle/CAS/tenant protections, updates the authoritative contract, and adds a state × conflict-result regression matrix.",
        ),
        (
            "Production closure",
            f"GO with caveat — exact runtime commit {DEPLOYED_COMMIT} serves 100% traffic on API revision {API_REVISION} and web revision {WEB_REVISION}; public health passed; supplied-tester Playwright passed 2/2 with cleanup; independent QA also passed both July 22 cases. Separate caveats: unlocked web dependency installation and two high npm-audit findings.",
        ),
    ]
    for decision in decisions:
        ws.append(decision)
    _table(ws, "ExecutiveDecisions", start, ws.max_row, 2)
    for row in range(start + 1, ws.max_row + 1):
        ws.row_dimensions[row].height = 64
    ws.freeze_panes = "A4"
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.oddFooter.center.text = "CaseOps | Ram 22 Jul 2026 | Confidential QA evidence"


def build_issue_assessment(wb: Workbook) -> None:
    ws = wb.create_sheet("Issue Assessment")
    _title(
        ws,
        "Issue assessment",
        "One source row was present. Classification separates whether the report is valid from whether it is a defect under the prior documented contract.",
        14,
    )
    headers = [
        "Issue ID",
        "Severity",
        "Reported Type",
        "Module",
        "Reported Status",
        "Evaluated Classification",
        "Validity",
        "Pre-change Verdict",
        "Requested Acceptance",
        "Root Cause",
        "Implemented Correction",
        "Adjacent Paths Audited",
        "Formal Post-work Verdict",
        "Closure Blocker",
    ]
    ws.append([])
    ws.append(headers)
    _header_row(ws, 4, len(headers))
    ws.append(
        [
            "BUG-001",
            "Medium",
            "Functional Bug",
            "Matter Management / Matter Status Update",
            "Open",
            "Valid enhancement",
            "Valid request; policy change",
            "Not fixed against the new 22-Jul acceptance",
            "Intake or On Hold may move to Active regardless of whether the latest conflict review is missing, pending, conflicted, cleared, waived, stale after reopen, or scoped to an older party name. Conflict review remains available and auditable.",
            "Previous July 15 work intentionally changed only direct creation. Existing-matter promotion retained a documented server gate and matching UI/tests/copy. The test suite therefore enforced a split contract instead of one cross-product invariant.",
            "Removed the activation gate/evaluator and gate-only UI. Retained advisory scans/resolution, tenant scoping, lifecycle provenance, row locking, optimistic concurrency, terminal-state protection, controlled reopen, and operational-child neutralisation.",
            "Direct create; Intake→Active; On Hold→Active; missing/pending/conflicted/cleared/waived/stale conflict results; changed opposing party; post-reopen activation; import-created Intake; cross-tenant review access; generic disposed writes; controlled reopen; UI edit/reload; public/current product copy.",
            "Properly fixed",
            f"None for BUG-001 — exact build identity ({DEPLOYED_COMMIT}), 100% production traffic, public health, committed production Playwright 2/2 with cleanup, and independent QA evidence are complete. The release's two pre-existing build/security caveats are tracked separately.",
        ]
    )
    _table(ws, "IssueAssessment", 4, 5, len(headers))
    ws.row_dimensions[5].height = 240
    _finish_sheet(
        ws,
        widths=[12, 11, 18, 30, 16, 22, 23, 28, 52, 58, 58, 70, 22, 55],
    )


def build_policy_matrix(wb: Workbook) -> None:
    ws = wb.create_sheet("Policy Matrix")
    _title(
        ws,
        "Permanent status and conflict-review contract",
        "This is the regression matrix—not a list of examples. Every row is an invariant that future changes must preserve unless the PRD is explicitly versioned again.",
        7,
    )
    headers = [
        "Entry Path",
        "Starting Status",
        "Conflict Review State",
        "Requested Operation",
        "Expected Result",
        "Safeguard That Remains",
        "Regression Coverage",
    ]
    ws.append([])
    ws.append(headers)
    _header_row(ws, 4, len(headers))
    rows = [
        ("Create", "n/a", "None", "Create directly as Active", "Allowed", "Create validation, tenant scope, audit", "Backend + July 15/22 E2E"),
        ("Edit", "Intake", "None", "PATCH status to Active", "Allowed", "CAS token and status consistency", "Backend + local + deployed production July 22 E2E passed (supplied tester and independent QA)"),
        ("Edit", "Intake", "Pending", "PATCH status to Active", "Allowed", "Review remains visible/auditable", "Backend matrix"),
        ("Edit", "Intake", "Conflicted", "PATCH status to Active", "Allowed", "Resolution capability still enforced", "Backend matrix"),
        ("Edit", "Intake", "Cleared / waived", "PATCH status to Active", "Allowed", "Review history preserved", "Backend matrix"),
        ("Edit", "On Hold", "None / any result", "PATCH status to Active", "Allowed", "CAS token and status consistency", "Backend state matrix covers On Hold → Active; local and deployed production July 22 E2E validate the shared nonblocking outcome through Intake/reopen paths"),
        ("Edit", "Intake", "Old party scope", "Change party and activate", "Allowed", "Old review remains historical", "Backend matrix"),
        ("Import", "Intake", "None", "PATCH imported matter to Active", "Allowed", "Import validation and tenant scope", "Backend regression"),
        ("Controlled reopen", "Disposed → Intake", "Pre-disposal result", "Activate without a new review", "Allowed", "Lifecycle version marks old review historical", "Backend lifecycle + local + deployed production July 22 E2E passed (dispose/reopen, historical clearance, reactivation, reload, cleanup)"),
        ("Generic edit", "Disposed", "Any", "PATCH status or operational fields", "Blocked", "Dedicated lifecycle endpoint, capability, reason", "Existing terminal lifecycle suite"),
        ("Conflict review", "Any operational status", "Any", "Run / list / resolve", "Allowed by capability", "Tenant isolation and audit", "Conflict-check API + UI regressions"),
        ("Cross tenant", "Any", "Another tenant's review", "Read / resolve", "404 / blocked", "Tenant isolation", "Backend regression"),
    ]
    for row in rows:
        ws.append(row)
    _table(ws, "PermanentPolicyMatrix", 4, ws.max_row, len(headers))
    for row in range(5, ws.max_row + 1):
        ws.row_dimensions[row].height = 48
    _finish_sheet(ws, widths=[24, 22, 26, 35, 22, 43, 38])


def build_verification_evidence(wb: Workbook) -> None:
    ws = wb.create_sheet("Verification Evidence")
    _title(
        ws,
        "Verification evidence",
        "A failed or blocked run remains visible. Local evidence cannot substitute for deployment proof, and unrelated setup failures are not counted as pass/fail evidence for BUG-001. The exact deployed evidence below now satisfies closure.",
        8,
    )
    headers = [
        "Layer",
        "Environment",
        "Command / Spec",
        "Acceptance Covered",
        "Result",
        "Evidence",
        "Counts / Timing",
        "Verdict Impact",
    ]
    ws.append([])
    ws.append(headers)
    _header_row(ws, 4, len(headers))
    rows = [
        (
            "Policy regression anchor",
            "Local API",
            "scripts/verify-backend.ps1 tests/test_conflict_checks.py -k test_intake_to_active_does_not_require_conflict_check",
            "Confirms current Intake→Active transition is nonblocking",
            "PASS",
            "The renamed regression returned HTTP 200, persisted Active, and emitted no conflict-gate denial audit.",
            "Included in the 59-test backend run",
            "Supports local candidate verification",
        ),
        (
            "Production baseline",
            "Production / Chromium",
            "ram-2026-07-15-prod.spec.ts --grep BUG-002|lifecycle",
            "Tester authentication and legacy lifecycle path",
            "BLOCKED (unrelated setup)",
            "Explicit tester sign-in succeeded, then the older test stopped because the production forum-state catalog remained empty. It never reached the lifecycle assertion and is not used as BUG-001 evidence.",
            "1 failed before target; 1 skipped",
            "No closure credit",
        ),
        (
            "Backend regression",
            "Local API",
            "scripts/verify-backend.ps1 tests/test_conflict_checks.py tests/test_matter_lifecycle.py tests/test_intake.py tests/test_matter_imports.py",
            "Optional-review state matrix; reopen history; import/intake path; retained safeguards",
            "PASS",
            "Canonical verifier passed Ruff across src/tests and every test in the four affected backend files.",
            "59 passed in 170.67s",
            "Local backend evidence complete",
        ),
        (
            "Frontend regression",
            "Local jsdom",
            "Focused Vitest; npm --prefix apps/web run typecheck; production Next.js build",
            "Status save without gate hint/error; conflict card distinguishes current from historical review",
            "PASS",
            "Three focused test files passed 19 tests, TypeScript passed, and the 64-route production build completed.",
            "19 tests; 3 files; typecheck + build pass",
            "Local frontend evidence complete",
        ),
        (
            "E2E harness reproducibility",
            "Fresh local checkout",
            "tests/e2e/global-setup.ts + scripts/verify-web.{ps1,sh}",
            "Fresh no-install-project venv import; build-time API origin/CSP consistency",
            "PASS after repair",
            "First runs exposed missing src-layout PYTHONPATH and a Next.js bundle baked with localhost while CSP allowed 127.0.0.1. The harness now sets both deterministically; no accidental editable install or stale bundle is required.",
            "Two setup defects reproduced and repaired",
            "Prevents false browser verdicts",
        ),
        (
            "Local browser E2E",
            "Fresh local tenant / Chromium",
            "tests/e2e/ram-2026-07-22-bugs.spec.ts",
            "Exact local identity; Intake/no-check; controlled dispose/reopen; historical clearance; final Active persistence",
            "PASS",
            "The shared legal/tester-email identity passed both July 22 cases: no-check Intake activation in 1.3s and cleared review → Active → Dispose → Reopen to Intake → Historical (stale) clearance → Active/read-back/reload in 2.1s. Combined July 15 + July 22 local execution passed 5/5. Password was runtime-only.",
            "5/5 combined in 20.5s; July 22: 2/2 (1.3s, 2.1s)",
            "Local candidate verified",
        ),
        (
            "Pre-deploy production browser E2E",
            "Production / Chromium",
            "tests/e2e/ram-2026-07-22-prod.spec.ts",
            "Committed Intake/no-check and controlled-reopen tester workflows with terminal cleanup",
            "FAIL — prior deployed policy reproduced",
            "Historical pre-deploy attempt: the extended spec authenticated and created a unique Intake matter. Its first activation returned HTTP 409 with the legacy Clear/Waived requirement; the second serial controlled-reopen case did not run. afterAll emitted no cleanup failure.",
            "1 failed at expected 200/actual 409; second serial case did not run",
            "Superseded by exact deployed-candidate PASS rows below",
        ),
        (
            "Production deployment identity",
            "Production / Cloud Run",
            f"scripts/deploy-prod.sh {DEPLOYED_COMMIT}; immutable revision/digest and health read-back",
            "Exact candidate identity; migration; scheduled jobs; 100% traffic; public health",
            "PASS",
            f"Runtime commit {DEPLOYED_COMMIT}; API {API_REVISION} at {API_DIGEST}; web {WEB_REVISION} at {WEB_DIGEST}; {MIGRATION_EXECUTION} passed 1/1; migrate plus four recurring jobs were pinned to the API digest; API health returned ok, web returned HTTP 200, and ClamAV was present.",
            "2 revisions at 100%; migration 1/1; 5 jobs pinned; health pass",
            "Exact candidate deployed and healthy",
        ),
        (
            "Supplied-tester production browser E2E",
            "Production / Chromium",
            "tests/e2e/ram-2026-07-22-prod.spec.ts",
            "Committed Intake/no-check and controlled-reopen/historical-clearance workflows with terminal cleanup",
            "PASS — exact deployed candidate",
            f"Using the supplied legal tester identity at runtime, both committed cases passed against deployed commit {DEPLOYED_COMMIT}; adjacent conflict review remained operational and afterAll cleanup passed. Password supplied at runtime; not stored in this workbook.",
            "2/2 in 71.6s (6.5s, 57.0s); cleanup passed",
            "Closes BUG-001 as Properly fixed",
        ),
        (
            "Independent QA production regression",
            "GitHub Actions / independent production QA tenant",
            f"GitHub run {INDEPENDENT_QA_RUN}",
            "Independent replay of both July 22 cases plus broader RAM and notice regressions",
            "PASS",
            f"Workflow checked out exact commit {DEPLOYED_COMMIT}. Both July 22 cases passed on an independent QA tenant; the RAM batch and notice module also passed. Four RAM skips were expected data-conditional legacy probes, not July 22 acceptance tests.",
            "July 22: 2/2 (8.9s, 14.4s); RAM: 46 passed + 4 expected conditional skips; notice: 2/2",
            "Independent corroboration of production closure",
        ),
    ]
    for row in rows:
        ws.append(row)
    _table(ws, "VerificationEvidence", 4, ws.max_row, len(headers))
    ws.conditional_formatting.add(
        f"E5:E{ws.max_row}",
        FormulaRule(formula=["ISNUMBER(SEARCH(\"PASS\",E5))"], fill=PatternFill("solid", fgColor=PALE_GREEN)),
    )
    ws.conditional_formatting.add(
        f"E5:E{ws.max_row}",
        FormulaRule(formula=["OR(ISNUMBER(SEARCH(\"FAIL\",E5)),ISNUMBER(SEARCH(\"BLOCK\",E5)))"], fill=PatternFill("solid", fgColor=PALE_RED)),
    )
    for row in range(5, ws.max_row + 1):
        ws.row_dimensions[row].height = 96
    for row in range(ws.max_row - 2, ws.max_row + 1):
        ws.row_dimensions[row].height = 120
    _finish_sheet(ws, widths=[24, 24, 65, 50, 31, 68, 24, 34])


def build_permanent_learnings(wb: Workbook) -> None:
    ws = wb.create_sheet("Permanent Learnings")
    _title(
        ws,
        "Brutal recurrence analysis and permanent rules",
        "These rules are mirrored in the repository bug-fixing skill and July 22 learning record so they survive beyond this workbook.",
        5,
    )
    headers = ["ID", "Where We Went Wrong", "Why Prior Checks Missed It", "Permanent Rule", "Regression Lock"]
    ws.append([])
    ws.append(headers)
    _header_row(ws, 4, len(headers))
    rows = [
        (
            "L1",
            "We treated direct Active creation and promotion of an existing Intake/On Hold matter as separate policies.",
            "July 15 acceptance and tests explicitly preserved the second gate, so the suite went green while users still encountered a mandatory check on another path into the same state.",
            "Define one invariant per business outcome and enumerate every entry path before implementation.",
            "Creation + Intake + On Hold + reopen + import rows in Policy Matrix.",
        ),
        (
            "L2",
            "The same policy was copied into backend decisions, UI error recognition, help text, marketing, PRD, and tests.",
            "A shallow service-only or copy-only patch would leave contradictory behavior and future developers would restore the gate from another ‘source of truth.’",
            "When policy changes, grep and update enforcement, recovery UI, copy, tests, PRD, task ledgers, and permanent skills in the same change.",
            "Repository-wide stale-policy scan must be clean except explicitly superseded history.",
        ),
        (
            "L3",
            "Regression tests encoded an implementation rule (‘latest check must be clear/waived’) instead of the user outcome (‘status may become Active’).",
            "Those tests actively prevented the requested behavior and made the obsolete rule look safe.",
            "Test externally observable outcomes with a state × review-result matrix; do not preserve obsolete internal decision objects.",
            "Missing, pending, conflicted, cleared, waived, stale, and changed-party cases all allow activation.",
        ),
        (
            "L4",
            "Reopen semantics conflated historical validity with operational permission.",
            "It is correct that a pre-disposal conflict result belongs to an older lifecycle, but incorrect under the new policy to make that provenance fact an activation blocker.",
            "Keep old results historical after reopen; never turn advisory provenance into an implicit gate.",
            "Reopen increments lifecycle version, preserves history, and permits immediate Intake→Active.",
        ),
        (
            "L5",
            "Prior closeout records drifted: one learning document held production evidence while authoritative task ledgers still said pending.",
            "Contradictory verdict sources make reopened reports hard to classify and encourage rework against stale assumptions.",
            "One formal verdict per issue; reconcile every authoritative ledger during closeout and retain explicit supersession notes for history.",
            f"July 15 records remain explicitly superseded; July 22 commit {DEPLOYED_COMMIT} was deployed, production-verified, and reconciled as Properly fixed.",
        ),
        (
            "L6",
            "We could have mistaken a friendlier 409 message or a review CTA for a fix.",
            "The user would still be unable to save Active, so the reported outcome would remain broken.",
            "A workflow bug is fixed only when the user completes the intended action on the deployed surface.",
            f"The committed production regression is tied to exact deployed SHA {DEPLOYED_COMMIT}; it asserts HTTP 2xx, dialog closure, reload persistence, and cleanup, and passed 2/2.",
        ),
    ]
    for row in rows:
        ws.append(row)
    _table(ws, "PermanentLearnings", 4, ws.max_row, len(headers))
    for row in range(5, ws.max_row + 1):
        ws.row_dimensions[row].height = 100
    _finish_sheet(ws, widths=[8, 58, 62, 62, 50])


def build(output: Path) -> None:
    wb = Workbook()
    build_executive_summary(wb)
    build_issue_assessment(wb)
    build_policy_matrix(wb)
    build_verification_evidence(wb)
    build_permanent_learnings(wb)
    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)

    # Read-back verification catches broken table ranges, invalid formulas, or
    # an accidentally truncated artifact before delivery.
    check = load_workbook(output, data_only=False, read_only=False)
    expected = [
        "Executive Summary",
        "Issue Assessment",
        "Policy Matrix",
        "Verification Evidence",
        "Permanent Learnings",
    ]
    if check.sheetnames != expected:
        raise RuntimeError(f"Unexpected sheet topology: {check.sheetnames!r}")
    if check["Issue Assessment"]["A5"].value != "BUG-001":
        raise RuntimeError("BUG-001 assessment row is missing")
    formal_verdict = str(check["Executive Summary"]["B10"].value or "")
    if "Properly fixed" not in formal_verdict or DEPLOYED_COMMIT not in formal_verdict:
        raise RuntimeError("Formal verdict is not tied to the exact deployed commit")
    if check["Issue Assessment"]["M5"].value != "Properly fixed":
        raise RuntimeError("BUG-001 issue verdict is not Properly fixed")

    verification = check["Verification Evidence"]
    production_passes = [
        row
        for row in verification.iter_rows(min_row=5, values_only=True)
        if row[4] and "PASS" in str(row[4]) and "production" in str(row[0]).lower()
    ]
    if len(production_passes) < 2:
        raise RuntimeError("Post-deploy production PASS evidence is incomplete")

    table_sheets = expected
    for sheet_name in table_sheets:
        sheet = check[sheet_name]
        if sheet.auto_filter.ref is not None:
            raise RuntimeError(f"Conflicting worksheet AutoFilter remains on {sheet_name}")
        if len(sheet.tables) != 1:
            raise RuntimeError(f"Expected exactly one Excel Table on {sheet_name}")

    formula_errors: list[str] = []
    unsafe_password_mentions: list[str] = []
    for sheet in check.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.data_type == "e":
                    formula_errors.append(f"{sheet.title}!{cell.coordinate}={cell.value}")
                if isinstance(cell.value, str) and "password" in cell.value.lower():
                    lowered = cell.value.lower()
                    if "not stored" not in lowered and "runtime-only" not in lowered:
                        unsafe_password_mentions.append(f"{sheet.title}!{cell.coordinate}")
    if formula_errors:
        raise RuntimeError(f"Formula errors found: {formula_errors}")
    if unsafe_password_mentions:
        raise RuntimeError(f"Unsafe password-bearing cells found: {unsafe_password_mentions}")
    if not check.calculation.fullCalcOnLoad or not check.calculation.forceFullCalc:
        raise RuntimeError("Workbook is not configured for full recalculation on open")
    check.close()
    print(f"wrote and verified {output}")


if __name__ == "__main__":
    target = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_OUTPUT
    build(target)
